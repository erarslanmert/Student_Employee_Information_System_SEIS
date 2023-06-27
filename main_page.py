# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'managerpage_studentdata.ui'
#
# Created by: PyQt5 UI code generator 5.9.2
#
# WARNING! All changes made in this file will be lost!

import os
import shutil
from collections import defaultdict
import cancel_class
import date_consistency
import delete_lesson
import numpy as np
import pandas as pd
import addition_options
import class_options
import datetime
import check_attandence
import connect_database
from datetime import timedelta
import change_data_options
import general_statistics_dashboard
import save_options
import json
import textwrap
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QDate, Qt, QTimer
from PyQt5.QtGui import QColor, QFont
from PyQt5.QtWidgets import QTableWidgetItem, QMessageBox, QAbstractItemView, QVBoxLayout, QGridLayout, QSpacerItem, \
    QSizePolicy, QHBoxLayout, QWidget, QHeaderView, QMenu, QAction, QStyle, QTableWidget, QComboBox
import data_objects
from PyQt5.QtGui import QTextCharFormat
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from dateutil.parser import parse
from dateutil.relativedelta import relativedelta

data_count_student = 0
data_count_teacher = 0
week_1 = 1
year_1 = 2023
firstdayofweek = (1,1,2023)
variable_1 = ' '
variable_to_change = ' '
cell_function_pointer = 0
cancel_program = 0
student_list = []
warning_pointer = 0
employee_full_list = []
personel_full_list = []
day_of_week = ['PAZARTESI','SALI','CARSAMBA','PERSEMBE','CUMA','CUMARTESI','PAZAR']

radio_button_selected = 0
group_class_list = ''
current_cell = []
date_list = []
attandence_list = []
attandence_checked = []

cell_clicked = []
schedule_list = []
restricted_student_list = []

saving_option = 0

lastly_selected_student = ''
lastly_selected_employee = ''

deleted_class = []
row_height = []
row_extender = 0
section_size = 0

class Ui_Dialog(object):
    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialog")
        Dialog.resize(1240, 888)
        Dialog.setStyleSheet("background-color: rgb(0, 81, 118);")
        Dialog.setWindowFlags(Dialog.windowFlags() | QtCore.Qt.WindowMaximizeButtonHint)
        Dialog.setWindowFlags(Dialog.windowFlags() | Qt.WindowMinimizeButtonHint)
        Dialog.setLocale(QtCore.QLocale(QtCore.QLocale.Turkish, QtCore.QLocale.Turkey))
        Dialog.setWindowIcon(QtGui.QIcon("logo_hq.png"))
        self.main_layout = QGridLayout(Dialog)
        self.frame_400 = QtWidgets.QFrame()
        self.frame_400.setGeometry(QtCore.QRect(0, 0, 321, 811))
        self.frame_400.setFixedWidth(321)
        self.frame_400.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_400.setObjectName("frame_400")
        self.pushButton_7 = QtWidgets.QPushButton(self.frame_400, clicked=lambda: change_data_options.open_change_options())
        self.pushButton_7.setGeometry(QtCore.QRect(50, 440, 231, 41))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(11)
        font.setItalic(False)
        self.pushButton_7.setFont(font)
        self.pushButton_7.setStyleSheet("background-color: rgb(255, 230, 207);")
        self.pushButton_7.setLocale(QtCore.QLocale(QtCore.QLocale.Turkish, QtCore.QLocale.Turkey))
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("management.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_7.setIcon(icon)
        self.pushButton_7.setIconSize(QtCore.QSize(30, 30))
        self.pushButton_7.setDefault(False)
        self.pushButton_7.setFlat(False)
        self.pushButton_7.setObjectName("pushButton_7")
        self.pushButton_3 = QtWidgets.QPushButton(self.frame_400, clicked = lambda : general_statistics_dashboard.barchart_general())
        self.pushButton_3.setGeometry(QtCore.QRect(50, 320, 231, 41))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(11)
        font.setItalic(False)
        self.pushButton_3.setFont(font)
        self.pushButton_3.setStyleSheet("background-color: rgb(255, 230, 207);")
        self.pushButton_3.setLocale(QtCore.QLocale(QtCore.QLocale.Turkish, QtCore.QLocale.Turkey))
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap("statistics.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_3.setIcon(icon1)
        self.pushButton_3.setIconSize(QtCore.QSize(30, 30))
        self.pushButton_3.setDefault(False)
        self.pushButton_3.setFlat(False)
        self.pushButton_3.setObjectName("pushButton_3")
        self.pushButton_8 = QtWidgets.QPushButton(Dialog, clicked = lambda : close_dialog())
        self.pushButton_8.setGeometry(QtCore.QRect(890, 820, 161, 41))
        self.pushButton_8.setFixedSize(161,41)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(11)
        font.setItalic(False)
        self.pushButton_8.setFont(font)
        self.pushButton_8.setStyleSheet("background-color: rgb(255, 230, 207);")
        self.pushButton_8.setLocale(QtCore.QLocale(QtCore.QLocale.Turkish, QtCore.QLocale.Turkey))
        icon2 = QtGui.QIcon()
        icon2.addPixmap(QtGui.QPixmap("log-out.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_8.setIcon(icon2)
        self.pushButton_8.setIconSize(QtCore.QSize(30, 30))
        self.pushButton_8.setDefault(False)
        self.pushButton_8.setFlat(False)
        self.pushButton_8.setObjectName("pushButton_8")
        self.label = QtWidgets.QLabel(self.frame_400)
        self.label.setGeometry(QtCore.QRect(100, 40, 121, 121))
        self.label.setFixedSize(121,121)
        self.label.setText("")
        self.label.setPixmap(QtGui.QPixmap("logo_hq.png"))
        self.label.setScaledContents(True)
        self.label.setObjectName("label")
        self.pushButton_2 = QtWidgets.QPushButton(self.frame_400, clicked = lambda : self.show_general_schedule())
        self.pushButton_2.setGeometry(QtCore.QRect(50, 260, 231, 41))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(11)
        font.setItalic(False)
        self.pushButton_2.setFont(font)
        self.pushButton_2.setStyleSheet("background-color: rgb(255, 230, 207);")
        self.pushButton_2.setLocale(QtCore.QLocale(QtCore.QLocale.Turkish, QtCore.QLocale.Turkey))
        icon3 = QtGui.QIcon()
        icon3.addPixmap(QtGui.QPixmap("schedule.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_2.setIcon(icon3)
        self.pushButton_2.setIconSize(QtCore.QSize(30, 30))
        self.pushButton_2.setDefault(False)
        self.pushButton_2.setFlat(False)
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_4 = QtWidgets.QPushButton(self.frame_400, clicked = lambda : addition_options.open_addition_selection())
        self.pushButton_4.setEnabled(True)
        self.pushButton_4.setGeometry(QtCore.QRect(50, 380, 231, 41))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(11)
        font.setItalic(False)
        self.pushButton_4.setFont(font)
        self.pushButton_4.setStyleSheet("background-color: rgb(255, 230, 207);")
        self.pushButton_4.setLocale(QtCore.QLocale(QtCore.QLocale.Turkish, QtCore.QLocale.Turkey))
        icon4 = QtGui.QIcon()
        icon4.addPixmap(QtGui.QPixmap("add.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_4.setIcon(icon4)
        self.pushButton_4.setIconSize(QtCore.QSize(30, 30))
        self.pushButton_4.setDefault(False)
        self.pushButton_4.setFlat(False)
        self.pushButton_4.setObjectName("pushButton_4")
        self.pushButton = QtWidgets.QPushButton(self.frame_400, clicked = lambda : self.show_data())
        self.pushButton.setGeometry(QtCore.QRect(50, 200, 231, 41))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(11)
        font.setItalic(False)
        self.pushButton.setFont(font)
        self.pushButton.setStyleSheet("background-color: rgb(255, 230, 207);")
        self.pushButton.setLocale(QtCore.QLocale(QtCore.QLocale.Turkish, QtCore.QLocale.Turkey))
        icon5 = QtGui.QIcon()
        icon5.addPixmap(QtGui.QPixmap("monitor.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton.setIcon(icon5)
        self.pushButton.setIconSize(QtCore.QSize(30, 30))
        self.pushButton.setDefault(False)
        self.pushButton.setFlat(False)
        self.pushButton.setObjectName("pushButton")
        self.pushButton_9 = QtWidgets.QPushButton(Dialog, clicked = lambda : self.save_file_dialog())
        self.pushButton_9.setGeometry(QtCore.QRect(720, 820, 161, 41))
        self.pushButton_9.setFixedSize(161,41)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(11)
        font.setItalic(False)
        self.pushButton_9.setFont(font)
        self.pushButton_9.setStyleSheet("background-color: rgb(255, 230, 207);")
        self.pushButton_9.setLocale(QtCore.QLocale(QtCore.QLocale.Turkish, QtCore.QLocale.Turkey))
        icon6 = QtGui.QIcon()
        icon6.addPixmap(QtGui.QPixmap("printer.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_9.setIcon(icon6)
        self.pushButton_9.setIconSize(QtCore.QSize(30, 30))
        self.pushButton_9.setDefault(False)
        self.pushButton_9.setFlat(False)
        self.pushButton_9.setObjectName("pushButton_9")
        self.calendarWidget = QtWidgets.QCalendarWidget(self.frame_400)
        self.calendarWidget.setEnabled(False)
        self.calendarWidget.setGeometry(QtCore.QRect(30, 500, 271, 301))
        font = QtGui.QFont()
        font.setPointSize(8)
        self.calendarWidget.setFont(font)
        self.calendarWidget.setContextMenuPolicy(QtCore.Qt.DefaultContextMenu)
        self.calendarWidget.setAutoFillBackground(False)
        self.calendarWidget.setStyleSheet("background-color: rgb(255, 254, 238);\n"
                                          "color: rgb(0, 85, 127);\n"
                                          "selection-background-color: rgb(211, 254, 255);\n"
                                          "selection-color: rgb(0, 0, 0);")
        self.calendarWidget.setGridVisible(False)
        self.calendarWidget.setSelectionMode(QtWidgets.QCalendarWidget.SingleSelection)
        self.calendarWidget.setHorizontalHeaderFormat(QtWidgets.QCalendarWidget.SingleLetterDayNames)
        self.calendarWidget.setVerticalHeaderFormat(QtWidgets.QCalendarWidget.NoVerticalHeader)
        self.calendarWidget.setNavigationBarVisible(True)
        self.calendarWidget.setDateEditEnabled(False)
        self.calendarWidget.setObjectName("calendarWidget")
        self.frame_2 = QtWidgets.QFrame(Dialog)
        self.frame_2.setGeometry(QtCore.QRect(310, 40, 911, 761))
        self.frame_2.setStyleSheet("color: rgb(255, 231, 213);\n"
                                   "background-color: rgb(0, 41, 61);")
        self.frame_2.setFrameShape(QtWidgets.QFrame.Box)
        self.frame_2.setFrameShadow(QtWidgets.QFrame.Plain)
        self.frame_2.setLineWidth(1)
        self.frame_2.setObjectName("frame_2")
        self.label_8 = QtWidgets.QLabel(self.frame_2)
        self.label_8.setGeometry(QtCore.QRect(10, 10, 151, 31))
        self.label_8.setFixedSize(151,31)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(11)
        font.setBold(True)
        font.setItalic(False)
        font.setWeight(75)
        self.label_8.setFont(font)
        self.label_8.setStyleSheet("color:  rgb(255, 230, 207);")
        self.label_8.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.label_8.setAlignment(QtCore.Qt.AlignLeading | QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        self.label_8.setObjectName("label_8")
        self.comboBox = QtWidgets.QComboBox(self.frame_2)
        self.comboBox.setGeometry(QtCore.QRect(260, 10, 271, 31))
        self.comboBox.setFixedSize(271,31)
        self.comboBox.setStyleSheet("background-color: rgb(249, 243, 255);\n"
                                    "color: rgb(0, 0, 0);")

        self.comboBox.setCurrentText("")
        self.comboBox.setObjectName("comboBox")
        font = QtGui.QFont()
        font.setPointSize(9)
        self.comboBox.setFont(font)
        self.frame_3 = QtWidgets.QFrame(self.frame_2)
        self.frame_3.setGeometry(QtCore.QRect(5, 95, 801, 661))
        self.frame_3.setStyleSheet("background-color: rgb(105, 100, 163);\n"
                                   "background-color: rgb(37, 67, 98);")
        self.frame_3.setFrameShape(QtWidgets.QFrame.Box)
        self.frame_3.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_3.setObjectName("frame_3")
        self.frame_301 = QtWidgets.QFrame(self.frame_3)
        self.frame_301.setGeometry(QtCore.QRect(320, 10, 361, 41))
        self.frame_301.setFixedSize(361,41)
        self.frame_301.setStyleSheet("background-color: rgb(105, 100, 163);\n"
                                     "background-color: rgb(37, 67, 98);")
        self.frame_301.setObjectName("frame_301")
        self.label_30 = QtWidgets.QLabel(self.frame_301)
        self.label_30.setGeometry(QtCore.QRect(90, 5, 191, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_30.setFont(font)
        self.label_30.setStyleSheet("color:  rgb(255, 230, 207);")
        self.label_30.setAlignment(QtCore.Qt.AlignLeading | QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        self.label_30.setObjectName("label_30")
        self.pushButton_6 = QtWidgets.QPushButton(self.frame_301, clicked = lambda:self.shift_data_student_right())
        self.pushButton_6.setGeometry(QtCore.QRect(290, 0, 31, 31))
        self.pushButton_6.setStyleSheet("background-color: rgb(213, 216, 255);\n"
                                        "color: rgb(0, 0, 0);")
        self.pushButton_6.setObjectName("pushButton_6")
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setItalic(False)
        self.pushButton_6.setFont(font)
        self.pushButton_10 = QtWidgets.QPushButton(self.frame_301, clicked = lambda:self.shift_data_student_left())
        self.pushButton_10.setGeometry(QtCore.QRect(40, 0, 31, 31))
        self.pushButton_10.setStyleSheet("background-color: rgb(213, 216, 255);\n"
                                         "color: rgb(0, 0, 0);")
        self.pushButton_10.setObjectName("pushButton_10")
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setItalic(False)
        self.pushButton_10.setFont(font)
        self.frame_302 = QtWidgets.QFrame(self.frame_3)
        self.frame_302.setGeometry(QtCore.QRect(30, 613, 350, 40))
        self.frame_302.setFixedSize(350,40)
        self.frame_302.setStyleSheet("background-color: rgb(105, 100, 163);\n"
                                   "background-color: rgb(37, 67, 98);")
        self.frame_302.setObjectName("frame_101")
        self.label_38 = QtWidgets.QLabel(self.frame_302)
        self.label_38.setGeometry(QtCore.QRect(100, 10, 25, 25))
        self.label_38.setText("")
        self.label_38.setPixmap(QtGui.QPixmap("yes.png"))
        self.label_38.setScaledContents(True)
        self.label_38.setObjectName("label_38")
        self.label_39 = QtWidgets.QLabel(self.frame_3)
        self.label_39.setGeometry(QtCore.QRect(320, 620, 161, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_39.setFont(font)
        self.label_39.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_39.setObjectName("label_39")
        self.label_40 = QtWidgets.QLabel(self.frame_3)
        self.label_40.setGeometry(QtCore.QRect(640, 620, 131, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_40.setFont(font)
        self.label_40.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_40.setObjectName("label_40")
        self.label_41 = QtWidgets.QLabel(self.frame_302)
        self.label_41.setGeometry(QtCore.QRect(10, 8, 400, 31))
        self.label_41.setFixedSize(450,31)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_41.setFont(font)
        self.label_41.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_41.setObjectName("label_41")
        self.frame_5 = QtWidgets.QFrame(self.frame_3)
        self.frame_5.setGeometry(QtCore.QRect(25, 60, 751, 361))
        self.frame_5.setMinimumHeight(361)
        self.frame_5.setStyleSheet("color: rgb(255, 230, 207);")
        self.frame_5.setFrameShape(QtWidgets.QFrame.Box)
        self.frame_5.setFrameShadow(QtWidgets.QFrame.Plain)
        self.frame_5.setObjectName("frame_5")
        self.label_44 = QtWidgets.QLabel(self.frame_5)
        self.label_44.setGeometry(QtCore.QRect(40, 179, 81, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_44.setFont(font)
        self.label_44.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_44.setObjectName("label_44")
        self.label_45 = QtWidgets.QLabel(self.frame_5)
        self.label_45.setGeometry(QtCore.QRect(40, 119, 61, 16))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_45.setFont(font)
        self.label_45.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_45.setObjectName("label_45")
        self.label_46 = QtWidgets.QLabel(self.frame_5)
        self.label_46.setGeometry(QtCore.QRect(40, 290, 91, 51))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_46.setFont(font)
        self.label_46.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_46.setWordWrap(True)
        self.label_46.setObjectName("label_46")
        self.label_47 = QtWidgets.QLabel(self.frame_5)
        self.label_47.setGeometry(QtCore.QRect(40, 149, 47, 16))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_47.setFont(font)
        self.label_47.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_47.setObjectName("label_47")
        self.label_48 = QtWidgets.QLabel(self.frame_5)
        self.label_48.setGeometry(QtCore.QRect(40, 209, 91, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_48.setFont(font)
        self.label_48.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_48.setObjectName("label_48")
        self.label_49 = QtWidgets.QLabel(self.frame_5)
        self.label_49.setGeometry(QtCore.QRect(400, 310, 91, 16))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_49.setFont(font)
        self.label_49.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_49.setObjectName("label_49")
        self.label_50 = QtWidgets.QLabel(self.frame_5)
        self.label_50.setGeometry(QtCore.QRect(40, 239, 91, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_50.setFont(font)
        self.label_50.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_50.setObjectName("label_50")
        self.label_51 = QtWidgets.QLabel(self.frame_5)
        self.label_51.setGeometry(QtCore.QRect(40, 269, 91, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_51.setFont(font)
        self.label_51.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_51.setObjectName("label_51")
        self.label_55 = QtWidgets.QLabel(self.frame_5)
        self.label_55.setGeometry(QtCore.QRect(400, 20, 61, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_55.setFont(font)
        self.label_55.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_55.setObjectName("label_55")
        self.label_56 = QtWidgets.QLabel(self.frame_5)
        self.label_56.setGeometry(QtCore.QRect(400, 50, 81, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_56.setFont(font)
        self.label_56.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_56.setObjectName("label_56")
        self.label_57 = QtWidgets.QLabel(self.frame_5)
        self.label_57.setGeometry(QtCore.QRect(400, 80, 91, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_57.setFont(font)
        self.label_57.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_57.setObjectName("label_57")
        self.label_58 = QtWidgets.QLabel(self.frame_5)
        self.label_58.setGeometry(QtCore.QRect(400, 100, 101, 51))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_58.setFont(font)
        self.label_58.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_58.setWordWrap(True)
        self.label_58.setObjectName("label_58")
        self.label_59 = QtWidgets.QLabel(self.frame_5)
        self.label_59.setGeometry(QtCore.QRect(400, 140, 101, 51))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_59.setFont(font)
        self.label_59.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_59.setWordWrap(True)
        self.label_59.setObjectName("label_59")
        self.label_60 = QtWidgets.QLabel(self.frame_5)
        self.label_60.setGeometry(QtCore.QRect(400, 210, 101, 51))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_60.setFont(font)
        self.label_60.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_60.setWordWrap(True)
        self.label_60.setObjectName("label_60")
        self.label_25 = QtWidgets.QLabel(self.frame_5)
        self.label_25.setGeometry(QtCore.QRect(200, 15, 81, 81))
        self.label_25.setFixedSize(81,81)
        self.label_25.setFrameShape(QtWidgets.QFrame.Box)
        self.label_25.setText("")
        self.label_25.setPixmap(QtGui.QPixmap("default-user-image.png"))
        self.label_25.setScaledContents(True)
        self.label_25.setObjectName("label_25")
        self.label_52 = QtWidgets.QLabel(self.frame_5)
        self.label_52.setGeometry(QtCore.QRect(140, 115, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_52.setFont(font)
        self.label_52.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                    "color: rgb(0, 0, 0);")
        self.label_52.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_52.setText("")
        self.label_52.setAlignment(QtCore.Qt.AlignCenter)
        self.label_52.setWordWrap(True)
        self.label_52.setObjectName("label_52")
        self.label_53 = QtWidgets.QLabel(self.frame_5)
        self.label_53.setGeometry(QtCore.QRect(140, 145, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_53.setFont(font)
        self.label_53.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                    "color: rgb(0, 0, 0);")
        self.label_53.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_53.setText("")
        self.label_53.setAlignment(QtCore.Qt.AlignCenter)
        self.label_53.setWordWrap(True)
        self.label_53.setObjectName("label_53")
        self.label_70 = QtWidgets.QLabel(self.frame_5)
        self.label_70.setGeometry(QtCore.QRect(140, 175, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_70.setFont(font)
        self.label_70.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                    "color: rgb(0, 0, 0);")
        self.label_70.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_70.setText("")
        self.label_70.setAlignment(QtCore.Qt.AlignCenter)
        self.label_70.setWordWrap(True)
        self.label_70.setObjectName("label_70")
        self.label_71 = QtWidgets.QLabel(self.frame_5)
        self.label_71.setGeometry(QtCore.QRect(140, 205, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_71.setFont(font)
        self.label_71.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                    "color: rgb(0, 0, 0);")
        self.label_71.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_71.setText("")
        self.label_71.setAlignment(QtCore.Qt.AlignCenter)
        self.label_71.setWordWrap(True)
        self.label_71.setObjectName("label_71")
        self.label_72 = QtWidgets.QLabel(self.frame_5)
        self.label_72.setGeometry(QtCore.QRect(140, 235, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_72.setFont(font)
        self.label_72.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                    "color: rgb(0, 0, 0);")
        self.label_72.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_72.setText("")
        self.label_72.setAlignment(QtCore.Qt.AlignCenter)
        self.label_72.setWordWrap(True)
        self.label_72.setObjectName("label_72")
        self.label_73 = QtWidgets.QLabel(self.frame_5)
        self.label_73.setGeometry(QtCore.QRect(140, 265, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_73.setFont(font)
        self.label_73.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                    "color: rgb(0, 0, 0);")
        self.label_73.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_73.setText("")
        self.label_73.setAlignment(QtCore.Qt.AlignCenter)
        self.label_73.setWordWrap(True)
        self.label_73.setObjectName("label_73")
        self.label_74 = QtWidgets.QLabel(self.frame_5)
        self.label_74.setGeometry(QtCore.QRect(140, 295, 211, 41))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_74.setFont(font)
        self.label_74.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                    "color: rgb(0, 0, 0);")
        self.label_74.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_74.setText("")
        self.label_74.setAlignment(QtCore.Qt.AlignCenter)
        self.label_74.setWordWrap(True)
        self.label_74.setObjectName("label_74")
        self.label_54 = QtWidgets.QLabel(self.frame_5)
        self.label_54.setGeometry(QtCore.QRect(500, 20, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_54.setFont(font)
        self.label_54.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                    "color: rgb(0, 0, 0);")
        self.label_54.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_54.setText("")
        self.label_54.setAlignment(QtCore.Qt.AlignCenter)
        self.label_54.setWordWrap(True)
        self.label_54.setObjectName("label_54")
        self.label_75 = QtWidgets.QLabel(self.frame_5)
        self.label_75.setGeometry(QtCore.QRect(500, 50, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_75.setFont(font)
        self.label_75.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                    "color: rgb(0, 0, 0);")
        self.label_75.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_75.setText("")
        self.label_75.setAlignment(QtCore.Qt.AlignCenter)
        self.label_75.setWordWrap(True)
        self.label_75.setObjectName("label_75")
        self.label_76 = QtWidgets.QLabel(self.frame_5)
        self.label_76.setGeometry(QtCore.QRect(500, 80, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_76.setFont(font)
        self.label_76.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                    "color: rgb(0, 0, 0);")
        self.label_76.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_76.setText("")
        self.label_76.setAlignment(QtCore.Qt.AlignCenter)
        self.label_76.setWordWrap(True)
        self.label_76.setObjectName("label_76")
        self.label_77 = QtWidgets.QLabel(self.frame_5)
        self.label_77.setGeometry(QtCore.QRect(500, 150, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_77.setFont(font)
        self.label_77.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                    "color: rgb(0, 0, 0);")
        self.label_77.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_77.setText("")
        self.label_77.setAlignment(QtCore.Qt.AlignCenter)
        self.label_77.setWordWrap(True)
        self.label_77.setObjectName("label_77")
        self.label_79 = QtWidgets.QLabel(self.frame_5)
        self.label_79.setGeometry(QtCore.QRect(500, 185, 211, 91))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_79.setFont(font)
        self.label_79.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                    "color: rgb(0, 0, 0);")
        self.label_79.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_79.setText("")
        self.label_79.setAlignment(QtCore.Qt.AlignCenter)
        self.label_79.setWordWrap(True)
        self.label_79.setObjectName("label_79")
        self.label_80 = QtWidgets.QLabel(self.frame_5)
        self.label_80.setGeometry(QtCore.QRect(500, 295, 211, 41))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_80.setFont(font)
        self.label_80.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                    "color: rgb(0, 0, 0);")
        self.label_80.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_80.setText("")
        self.label_80.setAlignment(QtCore.Qt.AlignCenter)
        self.label_80.setWordWrap(True)
        self.label_80.setObjectName("label_80")
        self.comboBox_5 = QtWidgets.QComboBox(self.frame_5)
        self.comboBox_5.setGeometry(QtCore.QRect(500, 115, 211, 21))
        self.comboBox_5.setFixedHeight(50)
        self.comboBox_5.setStyleSheet("color: rgb(0, 0, 0);\n"
                                      "background-color: rgb(255, 255, 255);")
        self.comboBox_5.setObjectName("comboBox_5")
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setItalic(False)
        self.comboBox_5.setFont(font)
        self.label_61 = QtWidgets.QLabel(self.frame_5)
        self.label_61.setGeometry(QtCore.QRect(40, 20, 61, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_61.setFont(font)
        self.label_61.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_61.setObjectName("label_61")
        self.label_64 = QtWidgets.QLabel(self.frame_5)
        self.label_64.setGeometry(QtCore.QRect(40, 80, 91, 16))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_64.setFont(font)
        self.label_64.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_64.setObjectName("label_64")
        self.label_65 = QtWidgets.QLabel(self.frame_5)
        self.label_65.setGeometry(QtCore.QRect(400, 80, 91, 16))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_65.setFont(font)
        self.label_65.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_65.setObjectName("label_65")
        self.label_68 = QtWidgets.QLabel(self.frame_5)
        self.label_68.setGeometry(QtCore.QRect(400, 20, 61, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_68.setFont(font)
        self.label_68.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_68.setObjectName("label_68")
        self.label_82 = QtWidgets.QLabel(self.frame_5)
        self.label_82.setGeometry(QtCore.QRect(140, 80, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_82.setFont(font)
        self.label_82.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                    "color: rgb(0, 0, 0);")
        self.label_82.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_82.setText("")
        self.label_82.setAlignment(QtCore.Qt.AlignCenter)
        self.label_82.setWordWrap(True)
        self.label_82.setObjectName("label_82")
        self.label_84 = QtWidgets.QLabel(self.frame_5)
        self.label_84.setGeometry(QtCore.QRect(140, 20, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_84.setFont(font)
        self.label_84.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                    "color: rgb(0, 0, 0);")
        self.label_84.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_84.setText("")
        self.label_84.setAlignment(QtCore.Qt.AlignCenter)
        self.label_84.setWordWrap(True)
        self.label_84.setObjectName("label_84")
        self.label_86 = QtWidgets.QLabel(self.frame_5)
        self.label_86.setGeometry(QtCore.QRect(500, 80, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_86.setFont(font)
        self.label_86.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                    "color: rgb(0, 0, 0);")
        self.label_86.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_86.setText("")
        self.label_86.setAlignment(QtCore.Qt.AlignCenter)
        self.label_86.setWordWrap(True)
        self.label_86.setObjectName("label_86")
        self.label_88 = QtWidgets.QLabel(self.frame_5)
        self.label_88.setGeometry(QtCore.QRect(500, 20, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_88.setFont(font)
        self.label_88.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                    "color: rgb(0, 0, 0);")
        self.label_88.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_88.setText("")
        self.label_88.setAlignment(QtCore.Qt.AlignCenter)
        self.label_88.setWordWrap(True)
        self.label_88.setObjectName("label_88")
        self.frame_4 = QtWidgets.QFrame(self.frame_3)
        self.frame_4.setGeometry(QtCore.QRect(10, 40, 781, 581))
        self.frame_4.setStyleSheet("color: rgb(255, 230, 207);")
        self.frame_4.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.frame_4.setFrameShadow(QtWidgets.QFrame.Plain)
        self.frame_4.setObjectName("frame_4")
        self.calendarWidget_2 = QtWidgets.QCalendarWidget(self.frame_4)
        self.calendarWidget_2.setGeometry(QtCore.QRect(10, 40, 761, 480))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(8)
        self.calendarWidget_2.setFont(font)
        self.calendarWidget_2.setAutoFillBackground(False)
        self.calendarWidget_2.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                            "selection-color: rgb(0, 85, 127);\n"
                                            "color: rgb(0, 85, 127);\n"
                                            "selection-background-color: rgb(211, 254, 255);")
        self.calendarWidget_2.setGridVisible(True)
        self.calendarWidget_2.setSelectionMode(QtWidgets.QCalendarWidget.SingleSelection)
        self.calendarWidget_2.setHorizontalHeaderFormat(QtWidgets.QCalendarWidget.LongDayNames)
        self.calendarWidget_2.setVerticalHeaderFormat(QtWidgets.QCalendarWidget.NoVerticalHeader)
        self.calendarWidget_2.setNavigationBarVisible(True)
        self.calendarWidget_2.setDateEditEnabled(True)
        self.calendarWidget_2.setObjectName("calendarWidget_2")
        self.tabWidget = QtWidgets.QTabWidget(self.frame_4)
        self.tabWidget.setGeometry(QtCore.QRect(10, 15, 216, 25))
        self.tabWidget.setFixedSize(216,25)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.tabWidget.setFont(font)
        self.tabWidget.setStyleSheet("color: rgb(0, 0, 0);\n"
                                     "background-color: rgb(243, 243, 243);")
        self.tabWidget.setTabPosition(QtWidgets.QTabWidget.North)
        self.tabWidget.setTabShape(QtWidgets.QTabWidget.Rounded)
        self.tabWidget.setElideMode(QtCore.Qt.ElideLeft)
        self.tabWidget.setObjectName("tabWidget")
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")
        self.tabWidget.addTab(self.tab, "")
        self.tab_2 = QtWidgets.QWidget()
        self.tab_2.setObjectName("tab_2")
        self.tabWidget.addTab(self.tab_2, "")
        self.frame_7 = QtWidgets.QFrame(self.frame_2)
        self.frame_7.setGeometry(QtCore.QRect(5, 95, 801, 661))
        self.frame_7.setStyleSheet("background-color: rgb(105, 100, 163);\n"
                                   "background-color: rgb(37, 67, 98);")
        self.frame_7.setFrameShape(QtWidgets.QFrame.Box)
        self.frame_7.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_7.setObjectName("frame_7")
        self.frame_401 = QtWidgets.QFrame(self.frame_7)
        self.frame_401.setGeometry(QtCore.QRect(320, 10, 361, 41))
        self.frame_401.setFixedSize(361, 41)
        self.frame_401.setStyleSheet("background-color: rgb(105, 100, 163);\n"
                                     "background-color: rgb(37, 67, 98);")
        self.frame_401.setObjectName("frame_401")
        self.label_31 = QtWidgets.QLabel(self.frame_401)
        self.label_31.setGeometry(QtCore.QRect(80, 5, 201, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_31.setFont(font)
        self.label_31.setStyleSheet("color:  rgb(255, 230, 207);")
        self.label_31.setAlignment(QtCore.Qt.AlignCenter)
        self.label_31.setObjectName("label_31")
        self.pushButton_11 = QtWidgets.QPushButton(self.frame_401, clicked = lambda:self.shift_data_teacher_right())
        self.pushButton_11.setGeometry(QtCore.QRect(290, 0, 31, 31))
        self.pushButton_11.setStyleSheet("background-color: rgb(213, 216, 255);\n"
                                         "color: rgb(0, 0, 0);")
        self.pushButton_11.setObjectName("pushButton_11")
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setItalic(False)
        self.pushButton_11.setFont(font)
        self.pushButton_12 = QtWidgets.QPushButton(self.frame_401, clicked = lambda:self.shift_data_teacher_left())
        self.pushButton_12.setGeometry(QtCore.QRect(40, 0, 31, 31))
        self.pushButton_12.setStyleSheet("background-color: rgb(213, 216, 255);\n"
                                         "color: rgb(0, 0, 0);")
        self.pushButton_12.setObjectName("pushButton_12")
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setItalic(False)
        self.pushButton_12.setFont(font)
        self.frame_402 = QtWidgets.QFrame(self.frame_3)
        self.frame_402.setGeometry(QtCore.QRect(30, 613, 400, 40))
        self.frame_402.setFixedSize(400, 40)
        self.frame_402.setStyleSheet("background-color: rgb(105, 100, 163);\n"
                                     "background-color: rgb(37, 67, 98);")
        self.frame_402.setObjectName("frame_402")
        self.label_43 = QtWidgets.QLabel(self.frame_7)
        self.label_43.setGeometry(QtCore.QRect(300, 620, 201, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_43.setFont(font)
        self.label_43.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_43.setObjectName("label_43")
        self.label_89 = QtWidgets.QLabel(self.frame_7)
        self.label_89.setGeometry(QtCore.QRect(600, 620, 171, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_89.setFont(font)
        self.label_89.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_89.setObjectName("label_89")
        self.label_90 = QtWidgets.QLabel(self.frame_402)
        self.label_90.setGeometry(QtCore.QRect(10, 8, 400, 31))
        self.label_90.setFixedSize(400,31)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_90.setFont(font)
        self.label_90.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_90.setObjectName("label_90")
        self.label_42 = QtWidgets.QLabel(self.frame_402)
        self.label_42.setGeometry(QtCore.QRect(100, 10, 25, 25))
        self.label_42.setText("")
        self.label_42.setPixmap(QtGui.QPixmap("yes.png"))
        self.label_42.setScaledContents(True)
        self.label_42.setObjectName("label_42")
        self.frame_8 = QtWidgets.QFrame(self.frame_7)
        self.frame_8.setGeometry(QtCore.QRect(25, 60, 751, 551))
        self.frame_8.setStyleSheet("color: rgb(255, 230, 207);")
        self.frame_8.setFrameShape(QtWidgets.QFrame.Box)
        self.frame_8.setFrameShadow(QtWidgets.QFrame.Plain)
        self.frame_8.setObjectName("frame_8")
        self.label_91 = QtWidgets.QLabel(self.frame_8)
        self.label_91.setGeometry(QtCore.QRect(40, 274, 81, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_91.setFont(font)
        self.label_91.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_91.setObjectName("label_91")
        self.label_92 = QtWidgets.QLabel(self.frame_8)
        self.label_92.setGeometry(QtCore.QRect(40, 204, 61, 16))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_92.setFont(font)
        self.label_92.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_92.setObjectName("label_92")
        self.label_93 = QtWidgets.QLabel(self.frame_8)
        self.label_93.setGeometry(QtCore.QRect(40, 400, 91, 51))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_93.setFont(font)
        self.label_93.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_93.setWordWrap(True)
        self.label_93.setObjectName("label_93")
        self.label_94 = QtWidgets.QLabel(self.frame_8)
        self.label_94.setGeometry(QtCore.QRect(40, 240, 47, 16))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_94.setFont(font)
        self.label_94.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_94.setObjectName("label_94")
        self.label_95 = QtWidgets.QLabel(self.frame_8)
        self.label_95.setGeometry(QtCore.QRect(40, 305, 91, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_95.setFont(font)
        self.label_95.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_95.setObjectName("label_95")
        self.label_97 = QtWidgets.QLabel(self.frame_8)
        self.label_97.setGeometry(QtCore.QRect(40, 330, 101, 41))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_97.setFont(font)
        self.label_97.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_97.setWordWrap(True)
        self.label_97.setObjectName("label_97")
        self.label_99 = QtWidgets.QLabel(self.frame_8)
        self.label_99.setGeometry(QtCore.QRect(400, 370, 81, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_99.setFont(font)
        self.label_99.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_99.setObjectName("label_99")
        self.label_100 = QtWidgets.QLabel(self.frame_8)
        self.label_100.setGeometry(QtCore.QRect(400, 400, 101, 41))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_100.setFont(font)
        self.label_100.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_100.setWordWrap(True)
        self.label_100.setObjectName("label_100")
        self.label_101 = QtWidgets.QLabel(self.frame_8)
        self.label_101.setGeometry(QtCore.QRect(400, 450, 91, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_101.setFont(font)
        self.label_101.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_101.setInputMethodHints(QtCore.Qt.ImhEmailCharactersOnly)
        self.label_101.setObjectName("label_101")
        self.label_26 = QtWidgets.QLabel(self.frame_8)
        self.label_26.setGeometry(QtCore.QRect(320, 30, 121, 121))
        self.label_26.setFixedSize(81,81)
        self.label_26.setFrameShape(QtWidgets.QFrame.Box)
        self.label_26.setText("")
        self.label_26.setPixmap(QtGui.QPixmap("default-user-image.png"))
        self.label_26.setScaledContents(True)
        self.label_26.setObjectName("label_26")
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setItalic(False)
        self.label_26.setFont(font)
        self.label_105 = QtWidgets.QLabel(self.frame_8)
        self.label_105.setGeometry(QtCore.QRect(140, 200, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_105.setFont(font)
        self.label_105.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_105.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_105.setText("")
        self.label_105.setAlignment(QtCore.Qt.AlignCenter)
        self.label_105.setWordWrap(True)
        self.label_105.setObjectName("label_105")
        self.label_106 = QtWidgets.QLabel(self.frame_8)
        self.label_106.setGeometry(QtCore.QRect(140, 235, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_106.setFont(font)
        self.label_106.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_106.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_106.setText("")
        self.label_106.setAlignment(QtCore.Qt.AlignCenter)
        self.label_106.setWordWrap(True)
        self.label_106.setObjectName("label_106")
        self.label_107 = QtWidgets.QLabel(self.frame_8)
        self.label_107.setGeometry(QtCore.QRect(140, 270, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_107.setFont(font)
        self.label_107.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_107.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_107.setText("")
        self.label_107.setAlignment(QtCore.Qt.AlignCenter)
        self.label_107.setWordWrap(True)
        self.label_107.setObjectName("label_107")
        self.label_108 = QtWidgets.QLabel(self.frame_8)
        self.label_108.setGeometry(QtCore.QRect(140, 305, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_108.setFont(font)
        self.label_108.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_108.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_108.setText("")
        self.label_108.setAlignment(QtCore.Qt.AlignCenter)
        self.label_108.setWordWrap(True)
        self.label_108.setObjectName("label_108")
        self.label_109 = QtWidgets.QLabel(self.frame_8)
        self.label_109.setGeometry(QtCore.QRect(140, 340, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_109.setFont(font)
        self.label_109.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_109.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_109.setText("")
        self.label_109.setAlignment(QtCore.Qt.AlignCenter)
        self.label_109.setWordWrap(True)
        self.label_109.setObjectName("label_109")
        self.label_112 = QtWidgets.QLabel(self.frame_8)
        self.label_112.setGeometry(QtCore.QRect(500, 370, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_112.setFont(font)
        self.label_112.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_112.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_112.setText("")
        self.label_112.setAlignment(QtCore.Qt.AlignCenter)
        self.label_112.setWordWrap(True)
        self.label_112.setObjectName("label_112")
        self.label_114 = QtWidgets.QLabel(self.frame_8)
        self.label_114.setGeometry(QtCore.QRect(500, 450, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_114.setFont(font)
        self.label_114.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_114.setInputMethodHints(QtCore.Qt.ImhEmailCharactersOnly)
        self.label_114.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_114.setText("")
        self.label_114.setAlignment(QtCore.Qt.AlignCenter)
        self.label_114.setWordWrap(True)
        self.label_114.setObjectName("label_114")
        self.label_130 = QtWidgets.QLabel(self.frame_8)
        self.label_130.setGeometry(QtCore.QRect(160, 500, 700, 50))
        self.label_130.setFixedSize(700,50)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setItalic(True)
        font.setWeight(60)
        self.label_130.setAlignment(QtCore.Qt.AlignCenter)
        self.label_130.setFont(font)
        self.label_130.setStyleSheet("color:  rgb(255, 230, 207);\n"
                                     "background-color: rgb(37, 67, 98);")
        self.label_130.setObjectName("label_130")
        self.label_96 = QtWidgets.QLabel(self.frame_8)
        self.label_96.setGeometry(QtCore.QRect(400, 240, 91, 16))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_96.setFont(font)
        self.label_96.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_96.setObjectName("label_96")
        self.label_117 = QtWidgets.QLabel(self.frame_8)
        self.label_117.setGeometry(QtCore.QRect(500, 230, 211, 41))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_117.setFont(font)
        self.label_117.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_117.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_117.setText("")
        self.label_117.setAlignment(QtCore.Qt.AlignCenter)
        self.label_117.setWordWrap(True)
        self.label_117.setObjectName("label_117")
        self.label_110 = QtWidgets.QLabel(self.frame_8)
        self.label_110.setGeometry(QtCore.QRect(500, 200, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_110.setFont(font)
        self.label_110.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_110.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_110.setText("")
        self.label_110.setAlignment(QtCore.Qt.AlignCenter)
        self.label_110.setWordWrap(True)
        self.label_110.setObjectName("label_110")
        self.label_98 = QtWidgets.QLabel(self.frame_8)
        self.label_98.setGeometry(QtCore.QRect(400, 200, 91, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_98.setFont(font)
        self.label_98.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_98.setObjectName("label_98")
        self.label_119 = QtWidgets.QLabel(self.frame_8)
        self.label_119.setGeometry(QtCore.QRect(400, 290, 91, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_119.setFont(font)
        self.label_119.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_119.setObjectName("label_119")
        self.label_120 = QtWidgets.QLabel(self.frame_8)
        self.label_120.setGeometry(QtCore.QRect(500, 290, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_120.setFont(font)
        self.label_120.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_120.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_120.setText("")
        self.label_120.setAlignment(QtCore.Qt.AlignCenter)
        self.label_120.setWordWrap(True)
        self.label_120.setObjectName("label_120")
        self.label_121 = QtWidgets.QLabel(self.frame_8)
        self.label_121.setGeometry(QtCore.QRect(400, 330, 61, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_121.setFont(font)
        self.label_121.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_121.setObjectName("label_121")
        self.label_122 = QtWidgets.QLabel(self.frame_8)
        self.label_122.setGeometry(QtCore.QRect(500, 330, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_122.setFont(font)
        self.label_122.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_122.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_122.setText("")
        self.label_122.setAlignment(QtCore.Qt.AlignCenter)
        self.label_122.setWordWrap(True)
        self.label_122.setObjectName("label_122")
        self.comboBox_9 = QtWidgets.QComboBox(self.frame_8)
        self.comboBox_9.setGeometry(QtCore.QRect(500, 410, 211, 21))
        self.comboBox_9.setFixedHeight(50)
        self.comboBox_9.setStyleSheet("color: rgb(0, 0, 0);\n"
                                      "background-color: rgb(255, 255, 255);")
        self.comboBox_9.setObjectName("comboBox_9")
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setItalic(False)
        self.comboBox_9.setFont(font)
        self.label_205 = QtWidgets.QLabel(self.frame_8)
        self.label_205.setGeometry(QtCore.QRect(140, 370, 211, 101))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_205.setFont(font)
        self.label_205.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_205.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_205.setText("")
        self.label_205.setAlignment(QtCore.Qt.AlignCenter)
        self.label_205.setWordWrap(True)
        self.label_205.setObjectName("label_205")
        self.frame_9 = QtWidgets.QFrame(self.frame_7)
        self.frame_9.setGeometry(QtCore.QRect(10, 40, 781, 581))
        self.frame_9.setStyleSheet("color: rgb(255, 230, 207);")
        self.frame_9.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.frame_9.setFrameShadow(QtWidgets.QFrame.Plain)
        self.frame_9.setObjectName("frame_9")
        self.label_200 = QtWidgets.QLabel(self.frame_9)
        self.label_200.setGeometry(QtCore.QRect(160, 521, 441, 41))
        self.label_200.setFixedSize(441, 41)
        self.label_200.setText("")
        self.label_200.setPixmap(QtGui.QPixmap("legend.png"))
        self.label_200.setScaledContents(True)
        self.label_200.setObjectName("label_200")
        self.label_201 = QtWidgets.QLabel(self.frame_4)
        self.label_201.setGeometry(QtCore.QRect(160, 521, 441, 41))
        self.label_201.setFixedSize(441,41)
        self.label_201.setText("")
        self.label_201.setPixmap(QtGui.QPixmap("legend.png"))
        self.label_201.setScaledContents(True)
        self.label_201.setObjectName("label_200")
        self.calendarWidget_4 = QtWidgets.QCalendarWidget(self.frame_9)
        self.calendarWidget_4.setGeometry(QtCore.QRect(10, 40, 761, 480))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(8)
        self.calendarWidget_4.setFont(font)
        self.calendarWidget_4.setAutoFillBackground(False)
        self.calendarWidget_4.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                            "selection-color: rgb(0, 85, 127);\n"
                                            "color: rgb(0, 85, 127);\n"
                                            "selection-background-color: rgb(211, 254, 255);")
        self.calendarWidget_4.setGridVisible(True)
        self.calendarWidget_4.setSelectionMode(QtWidgets.QCalendarWidget.SingleSelection)
        self.calendarWidget_4.setHorizontalHeaderFormat(QtWidgets.QCalendarWidget.LongDayNames)
        self.calendarWidget_4.setVerticalHeaderFormat(QtWidgets.QCalendarWidget.NoVerticalHeader)
        self.calendarWidget_4.setNavigationBarVisible(True)
        self.calendarWidget_4.setDateEditEnabled(True)
        self.calendarWidget_4.setObjectName("calendarWidget_4")
        self.tabWidget_3 = QtWidgets.QTabWidget(self.frame_9)
        self.tabWidget_3.setGeometry(QtCore.QRect(10, 15, 216, 25))
        self.tabWidget_3.setFixedSize(216,25)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.tabWidget_3.setFont(font)
        self.tabWidget_3.setStyleSheet("color: rgb(0, 0, 0);\n"
                                       "background-color: rgb(243, 243, 243);")
        self.tabWidget_3.setTabPosition(QtWidgets.QTabWidget.North)
        self.tabWidget_3.setTabShape(QtWidgets.QTabWidget.Rounded)
        self.tabWidget_3.setElideMode(QtCore.Qt.ElideLeft)
        self.tabWidget_3.setObjectName("tabWidget_3")
        self.tab_5 = QtWidgets.QWidget()
        self.tab_5.setObjectName("tab_5")
        self.tabWidget_3.addTab(self.tab_5, "")
        self.tab_6 = QtWidgets.QWidget()
        self.tab_6.setObjectName("tab_6")
        self.tabWidget_3.addTab(self.tab_6, "")
        self.frame_10 = QtWidgets.QFrame(self.frame_7)
        self.frame_10.setGeometry(QtCore.QRect(10, -610, 781, 581))
        self.frame_10.setStyleSheet("color: rgb(255, 230, 207);")
        self.frame_10.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.frame_10.setFrameShadow(QtWidgets.QFrame.Plain)
        self.frame_10.setObjectName("frame_10")
        self.calendarWidget_3 = QtWidgets.QCalendarWidget(self.frame_10)
        self.calendarWidget_3.setGeometry(QtCore.QRect(10, 40, 761, 531))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(8)
        self.calendarWidget_3.setFont(font)
        self.calendarWidget_3.setAutoFillBackground(False)
        self.calendarWidget_3.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                            "selection-color: rgb(0, 85, 127);\n"
                                            "color: rgb(0, 85, 127);\n"
                                            "selection-background-color: rgb(221, 255, 221);")
        self.calendarWidget_3.setGridVisible(True)
        self.calendarWidget_3.setSelectionMode(QtWidgets.QCalendarWidget.SingleSelection)
        self.calendarWidget_3.setHorizontalHeaderFormat(QtWidgets.QCalendarWidget.LongDayNames)
        self.calendarWidget_3.setVerticalHeaderFormat(QtWidgets.QCalendarWidget.NoVerticalHeader)
        self.calendarWidget_3.setNavigationBarVisible(True)
        self.calendarWidget_3.setDateEditEnabled(True)
        self.calendarWidget_3.setObjectName("calendarWidget_3")
        self.label_118 = QtWidgets.QLabel(self.frame_7)
        self.label_118.setGeometry(QtCore.QRect(40, 50, 121, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_118.setFont(font)
        self.label_118.setStyleSheet("color:  rgb(255, 230, 207);\n"
                                     "background-color: rgb(37, 67, 98);")
        self.label_118.setObjectName("label_118")
        self.label_118.hide()
        self.frame_11 = QtWidgets.QFrame(self.frame_2)
        self.frame_11.setGeometry(QtCore.QRect(5, 95, 801, 661))
        self.frame_11.setStyleSheet("background-color: rgb(105, 100, 163);\n"
                                    "background-color: rgb(37, 67, 98);")
        self.frame_11.setFrameShape(QtWidgets.QFrame.Box)
        self.frame_11.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_11.setObjectName("frame_11")
        self.label_32 = QtWidgets.QLabel(self.frame_11)
        self.label_32.setGeometry(QtCore.QRect(300, 15, 201, 21))
        self.label_32.setFixedSize(201,21)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_32.setFont(font)
        self.label_32.setStyleSheet("color:  rgb(255, 230, 207);")
        self.label_32.setAlignment(QtCore.Qt.AlignLeading | QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        self.label_32.setObjectName("label_32")
        self.frame_502 = QtWidgets.QFrame(self.frame_11)
        self.frame_502.setGeometry(QtCore.QRect(30, 613, 400, 40))
        self.frame_502.setFixedSize(400, 40)
        self.frame_502.setStyleSheet("background-color: rgb(105, 100, 163);\n"
                                     "background-color: rgb(37, 67, 98);")
        self.frame_502.setObjectName("frame_502")
        self.label_169 = QtWidgets.QLabel(self.frame_502)
        self.label_169.setGeometry(QtCore.QRect(10, 8, 400, 31))
        self.label_169.setFixedSize(400,31)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_169.setFont(font)
        self.label_169.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_169.setObjectName("label_169")
        self.label_113 = QtWidgets.QLabel(self.frame_502)
        self.label_113.setGeometry(QtCore.QRect(100, 10, 25, 255))
        self.label_113.setFixedSize(25,25)
        self.label_113.setText("")
        self.label_113.setPixmap(QtGui.QPixmap("yes.png"))
        self.label_113.setScaledContents(True)
        self.label_113.setObjectName("label_113")
        self.frame_12 = QtWidgets.QFrame(self.frame_11)
        self.frame_12.setGeometry(QtCore.QRect(25, 60, 751, 551))
        self.frame_12.setStyleSheet("color: rgb(255, 230, 207);")
        self.frame_12.setFrameShape(QtWidgets.QFrame.Box)
        self.frame_12.setFrameShadow(QtWidgets.QFrame.Plain)
        self.frame_12.setObjectName("frame_12")
        self.label_206 = QtWidgets.QLabel(self.frame_12)
        self.label_206.setGeometry(QtCore.QRect(140, 370, 211, 101))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_206.setFont(font)
        self.label_206.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_206.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_206.setText("")
        self.label_206.setAlignment(QtCore.Qt.AlignCenter)
        self.label_206.setWordWrap(True)
        self.label_206.setObjectName("label_206")
        self.label_170 = QtWidgets.QLabel(self.frame_12)
        self.label_170.setGeometry(QtCore.QRect(40, 264, 81, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_170.setFont(font)
        self.label_170.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_170.setObjectName("label_170")
        self.label_171 = QtWidgets.QLabel(self.frame_12)
        self.label_171.setGeometry(QtCore.QRect(40, 194, 61, 16))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_171.setFont(font)
        self.label_171.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_171.setObjectName("label_171")
        self.label_172 = QtWidgets.QLabel(self.frame_12)
        self.label_172.setGeometry(QtCore.QRect(40, 400, 91, 51))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_172.setFont(font)
        self.label_172.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_172.setWordWrap(True)
        self.label_172.setObjectName("label_172")
        self.label_173 = QtWidgets.QLabel(self.frame_12)
        self.label_173.setGeometry(QtCore.QRect(40, 230, 47, 16))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_173.setFont(font)
        self.label_173.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_173.setObjectName("label_173")
        self.label_174 = QtWidgets.QLabel(self.frame_12)
        self.label_174.setGeometry(QtCore.QRect(40, 295, 91, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_174.setFont(font)
        self.label_174.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_174.setObjectName("label_174")
        self.label_175 = QtWidgets.QLabel(self.frame_12)
        self.label_175.setGeometry(QtCore.QRect(40, 320, 101, 41))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_175.setFont(font)
        self.label_175.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_175.setWordWrap(True)
        self.label_175.setObjectName("label_175")
        self.label_176 = QtWidgets.QLabel(self.frame_12)
        self.label_176.setGeometry(QtCore.QRect(400, 370, 81, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_176.setFont(font)
        self.label_176.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_176.setObjectName("label_176")
        self.label_177 = QtWidgets.QLabel(self.frame_12)
        self.label_177.setGeometry(QtCore.QRect(400, 400, 101, 41))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_177.setFont(font)
        self.label_177.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_177.setWordWrap(True)
        self.label_177.setObjectName("label_177")
        self.label_178 = QtWidgets.QLabel(self.frame_12)
        self.label_178.setGeometry(QtCore.QRect(400, 450, 91, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_178.setFont(font)
        self.label_178.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_178.setInputMethodHints(QtCore.Qt.ImhEmailCharactersOnly)
        self.label_178.setObjectName("label_178")
        self.label_29 = QtWidgets.QLabel(self.frame_12)
        self.label_29.setGeometry(QtCore.QRect(320, 30, 121, 121))
        self.label_29.setFixedSize(81,81)
        self.label_29.setFrameShape(QtWidgets.QFrame.Box)
        self.label_29.setText("")
        self.label_29.setPixmap(QtGui.QPixmap("default-user-image.png"))
        self.label_29.setScaledContents(True)
        self.label_29.setObjectName("label_29")
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setItalic(False)
        self.label_29.setFont(font)
        self.label_179 = QtWidgets.QLabel(self.frame_12)
        self.label_179.setGeometry(QtCore.QRect(140, 190, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_179.setFont(font)
        self.label_179.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_179.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_179.setText("")
        self.label_179.setAlignment(QtCore.Qt.AlignCenter)
        self.label_179.setWordWrap(True)
        self.label_179.setObjectName("label_179")
        self.label_180 = QtWidgets.QLabel(self.frame_12)
        self.label_180.setGeometry(QtCore.QRect(140, 225, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_180.setFont(font)
        self.label_180.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_180.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_180.setText("")
        self.label_180.setAlignment(QtCore.Qt.AlignCenter)
        self.label_180.setWordWrap(True)
        self.label_180.setObjectName("label_180")
        self.label_181 = QtWidgets.QLabel(self.frame_12)
        self.label_181.setGeometry(QtCore.QRect(140, 260, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_181.setFont(font)
        self.label_181.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_181.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_181.setText("")
        self.label_181.setAlignment(QtCore.Qt.AlignCenter)
        self.label_181.setWordWrap(True)
        self.label_181.setObjectName("label_181")
        self.label_182 = QtWidgets.QLabel(self.frame_12)
        self.label_182.setGeometry(QtCore.QRect(140, 295, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_182.setFont(font)
        self.label_182.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_182.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_182.setText("")
        self.label_182.setAlignment(QtCore.Qt.AlignCenter)
        self.label_182.setWordWrap(True)
        self.label_182.setObjectName("label_182")
        self.label_131 = QtWidgets.QLabel(self.frame_12)
        self.label_131.setGeometry(QtCore.QRect(160, 500, 451, 20))
        self.label_131.setFixedSize(700,50)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setItalic(True)
        font.setWeight(60)
        self.label_131.setAlignment(QtCore.Qt.AlignCenter)
        self.label_131.setFont(font)
        self.label_131.setStyleSheet("color:  rgb(255, 230, 207);\n"
                                     "background-color: rgb(37, 67, 98);")
        self.label_131.setObjectName("label_131")
        self.label_183 = QtWidgets.QLabel(self.frame_12)
        self.label_183.setGeometry(QtCore.QRect(140, 330, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_183.setFont(font)
        self.label_183.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_183.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_183.setText("")
        self.label_183.setAlignment(QtCore.Qt.AlignCenter)
        self.label_183.setWordWrap(True)
        self.label_183.setObjectName("label_183")
        self.label_184 = QtWidgets.QLabel(self.frame_12)
        self.label_184.setGeometry(QtCore.QRect(500, 370, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_184.setFont(font)
        self.label_184.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_184.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_184.setText("")
        self.label_184.setAlignment(QtCore.Qt.AlignCenter)
        self.label_184.setWordWrap(True)
        self.label_184.setObjectName("label_184")
        self.label_185 = QtWidgets.QLabel(self.frame_12)
        self.label_185.setGeometry(QtCore.QRect(500, 450, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_185.setFont(font)
        self.label_185.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_185.setInputMethodHints(QtCore.Qt.ImhEmailCharactersOnly)
        self.label_185.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_185.setText("")
        self.label_185.setAlignment(QtCore.Qt.AlignCenter)
        self.label_185.setWordWrap(True)
        self.label_185.setObjectName("label_185")
        self.label_186 = QtWidgets.QLabel(self.frame_12)
        self.label_186.setGeometry(QtCore.QRect(400, 240, 91, 16))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_186.setFont(font)
        self.label_186.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_186.setObjectName("label_186")
        self.label_187 = QtWidgets.QLabel(self.frame_12)
        self.label_187.setGeometry(QtCore.QRect(500, 230, 211, 41))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_187.setFont(font)
        self.label_187.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_187.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_187.setText("")
        self.label_187.setAlignment(QtCore.Qt.AlignCenter)
        self.label_187.setWordWrap(True)
        self.label_187.setObjectName("label_187")
        self.label_188 = QtWidgets.QLabel(self.frame_12)
        self.label_188.setGeometry(QtCore.QRect(500, 190, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_188.setFont(font)
        self.label_188.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_188.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_188.setText("")
        self.label_188.setAlignment(QtCore.Qt.AlignCenter)
        self.label_188.setWordWrap(True)
        self.label_188.setObjectName("label_188")
        self.label_189 = QtWidgets.QLabel(self.frame_12)
        self.label_189.setGeometry(QtCore.QRect(400, 190, 91, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_189.setFont(font)
        self.label_189.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_189.setObjectName("label_189")
        self.label_190 = QtWidgets.QLabel(self.frame_12)
        self.label_190.setGeometry(QtCore.QRect(400, 290, 91, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_190.setFont(font)
        self.label_190.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_190.setObjectName("label_190")
        self.label_191 = QtWidgets.QLabel(self.frame_12)
        self.label_191.setGeometry(QtCore.QRect(500, 290, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_191.setFont(font)
        self.label_191.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_191.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_191.setText("")
        self.label_191.setAlignment(QtCore.Qt.AlignCenter)
        self.label_191.setWordWrap(True)
        self.label_191.setObjectName("label_191")
        self.label_192 = QtWidgets.QLabel(self.frame_12)
        self.label_192.setGeometry(QtCore.QRect(400, 330, 61, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        font.setWeight(75)
        self.label_192.setFont(font)
        self.label_192.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_192.setObjectName("label_192")
        self.label_193 = QtWidgets.QLabel(self.frame_12)
        self.label_193.setGeometry(QtCore.QRect(500, 330, 211, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        self.label_193.setFont(font)
        self.label_193.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                     "color: rgb(0, 0, 0);")
        self.label_193.setFrameShape(QtWidgets.QFrame.Panel)
        self.label_193.setText("")
        self.label_193.setAlignment(QtCore.Qt.AlignCenter)
        self.label_193.setWordWrap(True)
        self.label_193.setObjectName("label_193")
        self.comboBox_10 = QtWidgets.QComboBox(self.frame_12)
        self.comboBox_10.setGeometry(QtCore.QRect(500, 410, 211, 21))
        self.comboBox_10.setFixedHeight(50)
        self.comboBox_10.setStyleSheet("color: rgb(0, 0, 0);\n"
                                       "background-color: rgb(255, 255, 255);")
        self.comboBox_10.setObjectName("comboBox_10")
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setItalic(False)
        self.comboBox_10.setFont(font)
        self.label_2 = QtWidgets.QLabel(self.frame_12)
        self.label_2.setGeometry(QtCore.QRect(140, 370, 211, 101))
        self.label_2.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                      "color: rgb(0, 0, 0);")
        self.label_2.setObjectName("label_2")
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setItalic(False)
        self.label_2.setFont(font)
        self.frame_14 = QtWidgets.QFrame(self.frame_11)
        self.frame_14.setGeometry(QtCore.QRect(10, -610, 781, 581))
        self.frame_14.setStyleSheet("color: rgb(255, 230, 207);")
        self.frame_14.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.frame_14.setFrameShadow(QtWidgets.QFrame.Plain)
        self.frame_14.setObjectName("frame_14")
        self.calendarWidget_6 = QtWidgets.QCalendarWidget(self.frame_14)
        self.calendarWidget_6.setGeometry(QtCore.QRect(10, 40, 761, 531))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(8)
        self.calendarWidget_6.setFont(font)
        self.calendarWidget_6.setAutoFillBackground(False)
        self.calendarWidget_6.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                                            "selection-color: rgb(0, 85, 127);\n"
                                            "color: rgb(0, 85, 127);\n"
                                            "selection-background-color: rgb(221, 255, 221);")
        self.calendarWidget_6.setGridVisible(True)
        self.calendarWidget_6.setSelectionMode(QtWidgets.QCalendarWidget.SingleSelection)
        self.calendarWidget_6.setHorizontalHeaderFormat(QtWidgets.QCalendarWidget.LongDayNames)
        self.calendarWidget_6.setVerticalHeaderFormat(QtWidgets.QCalendarWidget.NoVerticalHeader)
        self.calendarWidget_6.setNavigationBarVisible(True)
        self.calendarWidget_6.setDateEditEnabled(True)
        self.calendarWidget_6.setObjectName("calendarWidget_6")
        self.label_194 = QtWidgets.QLabel(self.frame_11)
        self.label_194.setGeometry(QtCore.QRect(40, 50, 121, 21))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(True)
        self.label_194.setFont(font)
        self.label_194.setStyleSheet("color:  rgb(255, 230, 207);\n"
                                     "background-color: rgb(37, 67, 98);")
        self.label_194.setObjectName("label_194")
        self.comboBox_2 = QtWidgets.QComboBox(self.frame_2)
        self.comboBox_2.setGeometry(QtCore.QRect(260, 50, 271, 31))
        self.comboBox_2.setFixedSize(271,31)
        self.comboBox_2.setStyleSheet("background-color: rgb(249, 243, 255);\n"
                                      "color: rgb(0, 0, 0);")

        self.comboBox_2.setObjectName("comboBox_2")
        self.comboBox_2.setMinimumContentsLength(1)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setItalic(False)
        self.comboBox_2.setFont(font)

        self.comboBox_3 = QtWidgets.QComboBox(self.frame_2)
        self.comboBox_3.setGeometry(QtCore.QRect(260, 50, 271, 31))
        self.comboBox_3.setStyleSheet("background-color: rgb(249, 243, 255);\n"
                                      "color: rgb(0, 0, 0);")

        self.comboBox_3.setObjectName("comboBox_2")
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setItalic(False)
        self.comboBox_3.setFont(font)
        icon7 = QtGui.QIcon()
        icon7.addPixmap(QtGui.QPixmap("211818_search_icon.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.frame_11.hide()
        self.frame_7.hide()
        self.frame_4.hide()
        self.frame_9.hide()


        self.frame_13 = QtWidgets.QFrame(Dialog)
        self.frame_13.setGeometry(QtCore.QRect(310, 40, 811, 761))
        self.frame_13.setStyleSheet("color: rgb(255, 231, 213);\n"
                                    "background-color: rgb(0, 41, 61);")
        self.frame_13.setFrameShape(QtWidgets.QFrame.Box)
        self.frame_13.setFrameShadow(QtWidgets.QFrame.Plain)
        self.frame_13.setLineWidth(1)
        self.frame_13.setObjectName("frame_13")
        self.label_10 = QtWidgets.QLabel(self.frame_13)
        self.label_10.setGeometry(QtCore.QRect(10, 10, 151, 31))
        self.label_10.setFixedSize(151,31)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(11)
        font.setBold(True)
        font.setItalic(False)
        font.setWeight(75)
        self.label_10.setFont(font)
        self.label_10.setStyleSheet("color:  rgb(255, 230, 207);")
        self.label_10.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.label_10.setAlignment(QtCore.Qt.AlignLeading | QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        self.label_10.setObjectName("label_10")
        self.pushButton_14 = QtWidgets.QPushButton(self.frame_13, clicked = lambda : self.copy_schedule_next())
        self.pushButton_14.setGeometry(QtCore.QRect(20, 43, 121, 25))
        self.pushButton_14.setFixedSize(121,25)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setItalic(False)
        self.pushButton_14.setFont(font)
        self.pushButton_14.setStyleSheet("color: rgb(0, 0, 0);\n"
                                         "background-color: rgb(255, 230, 207);")

        self.pushButton_14.setLocale(QtCore.QLocale(QtCore.QLocale.Turkish, QtCore.QLocale.Turkey))
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap("copy.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_14.setIcon(icon1)
        self.pushButton_14.setIconSize(QtCore.QSize(15,15))
        self.pushButton_14.setDefault(False)
        self.pushButton_14.setFlat(False)
        self.pushButton_14.setObjectName("pushButton_14")
        self.pushButton_100 = QtWidgets.QPushButton(self.frame_13, clicked=lambda: check_maximized())
        self.pushButton_100.setFixedSize(40, 20)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(11)
        font.setItalic(False)
        self.pushButton_100.setFont(font)
        self.pushButton_100.setStyleSheet("background-color: rgb(255, 230, 207);")
        self.pushButton_100.setLocale(QtCore.QLocale(QtCore.QLocale.Turkish, QtCore.QLocale.Turkey))
        icon5 = QtGui.QIcon()
        icon5.addPixmap(QtGui.QPixmap("expand.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_100.setIcon(icon5)
        self.pushButton_100.setIconSize(QtCore.QSize(15, 15))
        self.pushButton_100.setDefault(False)
        self.pushButton_100.setFlat(False)
        self.pushButton_100.setObjectName("pushButton_100")
        self.tableWidget = QtWidgets.QTableWidget(self.frame_13)
        self.tableWidget.setGeometry(QtCore.QRect(10, 90, 791, 661))
        self.tableWidget.setStyleSheet("color: rgb(0, 0, 0);\n"
                                       "background-color: rgb(255, 250, 239);")
        self.tableWidget.setGridStyle(QtCore.Qt.SolidLine)
        self.tableWidget.setRowCount(70)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.horizontalHeader().setCascadingSectionResizes(False)
        self.tableWidget.horizontalHeader().setDefaultSectionSize(300)
        self.tableWidget.horizontalHeader().setMinimumSectionSize(200)
        self.tableWidget.verticalHeader().setMinimumSectionSize(50)
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        font = QtGui.QFont()
        font.setPointSize(9)
        self.tableWidget.verticalHeader().setFont(font)
        self.tableWidget.horizontalHeader().setFont(font)
        self.tableWidget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget.resizeRowsToContents()
        self.pushButton_19 = QtWidgets.QPushButton(self.frame_13, clicked = lambda : self.step_back_weekly())
        self.pushButton_19.setGeometry(QtCore.QRect(270, 40, 31, 31))
        self.pushButton_19.setFixedSize(31,31)
        self.pushButton_19.setStyleSheet("background-color: rgb(213, 216, 255);\n"
                                         "color: rgb(0, 0, 0);")
        self.pushButton_19.setObjectName("pushButton_19")
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setItalic(False)
        self.pushButton_19.setFont(font)
        self.pushButton_20 = QtWidgets.QPushButton(self.frame_13, clicked = lambda : self.step_fwd_weekly())
        self.pushButton_20.setGeometry(QtCore.QRect(520, 40, 31, 31))
        self.pushButton_20.setFixedSize(31,31)
        self.pushButton_20.setStyleSheet("background-color: rgb(213, 216, 255);\n"
                                         "color: rgb(0, 0, 0);")
        self.pushButton_20.setObjectName("pushButton_20")
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setItalic(False)
        self.pushButton_20.setFont(font)
        self.label_35 = QtWidgets.QLabel(self.frame_13)
        self.label_35.setGeometry(QtCore.QRect(315, 45, 191, 21))
        self.label_35.setFixedSize(191,21)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_35.setFont(font)
        self.label_35.setStyleSheet("color:  rgb(255, 230, 207);")
        self.label_35.setAlignment(QtCore.Qt.AlignCenter)
        self.label_35.setObjectName("label_35")
        self.dateEdit = QtWidgets.QDateEdit(self.frame_13)
        self.dateEdit.setGeometry(QtCore.QRect(610, 43, 171, 22))
        self.dateEdit.setFixedSize(171,22)
        self.dateEdit.setAlignment(QtCore.Qt.AlignCenter)
        self.dateEdit.setCurrentSection(QtWidgets.QDateTimeEdit.DaySection)
        self.dateEdit.setDate(QtCore.QDate(2023, 1, 1))
        self.dateEdit.setObjectName("dateEdit")
        self.dateEdit.setStyleSheet("background-color: rgb(255, 254, 238);\n"
                                          "color: rgb(0, 0, 0);")
        self.dateEdit.setDate(QDate.currentDate())
        self.dateEdit.setDisabled(True)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setItalic(False)
        self.dateEdit.setFont(font)
        self.tabWidget.setCurrentIndex(0)
        self.tabWidget.addTab(self.tab_2, "")
        self.tableWidget_2 = QtWidgets.QTableWidget(self.frame_4)
        self.tableWidget_2.setGeometry(QtCore.QRect(10, 40, 761, 480))
        self.tableWidget_2.setStyleSheet("color: rgb(0, 0, 0);\n"
                                         "background-color: rgb(255, 250, 239);")
        self.tableWidget_2.setGridStyle(QtCore.Qt.SolidLine)
        self.tableWidget_2.setRowCount(9)
        self.tableWidget_2.setColumnCount(7)
        self.tableWidget_2.setObjectName("tableWidget_2")
        font = QtGui.QFont()
        font.setPointSize(9)
        self.tableWidget_2.setFont(font)
        self.tableWidget_2.verticalHeader().setFont(font)
        self.tableWidget_2.horizontalHeader().setFont(font)
        self.tableWidget_2.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tableWidget_2.hide()
        self.tableWidget_2.setSizePolicy(QSizePolicy.Expanding,QSizePolicy.Expanding)
        self.tableWidget_2.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget_2.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget_2.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.spinBox = QtWidgets.QSpinBox(self.frame_4)
        self.spinBox.setGeometry(QtCore.QRect(570, 10, 91, 22))
        self.spinBox.setFixedSize(91,22)
        self.spinBox.setStyleSheet("background-color: rgb(255, 255, 255);;\n"
                                          "color: rgb(0, 0, 0);")
        self.spinBox.setAlignment(QtCore.Qt.AlignCenter)
        self.spinBox.setMinimum(2020)
        self.spinBox.setMaximum(2100)
        self.spinBox.setObjectName("spinBox")
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setItalic(False)
        self.spinBox.setFont(font)
        self.spinBox_2 = QtWidgets.QSpinBox(self.frame_4)
        self.spinBox_2.setGeometry(QtCore.QRect(670, 10, 91, 22))
        self.spinBox_2.setFixedSize(91,22)
        self.spinBox_2.setStyleSheet("background-color: rgb(255, 255, 255);;\n"
                                          "color: rgb(0, 0, 0);")
        self.spinBox_2.setAlignment(QtCore.Qt.AlignCenter)
        self.spinBox_2.setMinimum(1)
        self.spinBox_2.setMaximum(52)
        self.spinBox_2.setObjectName("spinBox_2")
        self.spinBox_2.setValue(QDate.currentDate().weekNumber()[0])
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setItalic(False)
        self.spinBox_2.setFont(font)
        self.spinBox.setValue(QDate.currentDate().weekNumber()[1])
        self.calendarWidget_2.setSelectionMode(QtWidgets.QCalendarWidget.NoSelection)

        self.tableWidget_3 = QtWidgets.QTableWidget(self.frame_9)
        self.tableWidget_3.setGeometry(QtCore.QRect(10, 40, 761, 480))
        self.tableWidget_3.setStyleSheet("color: rgb(0, 0, 0);\n"
                                         "background-color: rgb(255, 250, 239);")
        self.tableWidget_3.setGridStyle(QtCore.Qt.SolidLine)
        self.tableWidget_3.setRowCount(9)
        self.tableWidget_3.setColumnCount(7)
        self.tableWidget_3.setObjectName("tableWidget_3")
        self.tableWidget_3.horizontalHeader().setCascadingSectionResizes(False)
        self.tableWidget_3.horizontalHeader().setDefaultSectionSize(98)
        self.tableWidget_3.verticalHeader().setDefaultSectionSize(52)
        self.tableWidget_3.horizontalHeader().setMinimumSectionSize(50)
        self.tableWidget_3.verticalHeader().setMinimumSectionSize(50)
        font = QtGui.QFont()
        font.setPointSize(9)
        self.tableWidget_3.verticalHeader().setFont(font)
        self.tableWidget_3.horizontalHeader().setFont(font)
        self.tableWidget_3.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tableWidget_3.hide()
        self.tableWidget_3.setSizePolicy(QSizePolicy.Expanding,QSizePolicy.Expanding)
        self.tableWidget_3.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget_3.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget_3.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.spinBox_3 = QtWidgets.QSpinBox(self.frame_9)
        self.spinBox_3.setGeometry(QtCore.QRect(570, 10, 91, 22))
        self.spinBox_3.setStyleSheet("background-color: rgb(255, 255, 255);;\n"
                                          "color: rgb(0, 0, 0);")
        self.spinBox_3.setAlignment(QtCore.Qt.AlignCenter)
        self.spinBox_3.setMinimum(2020)
        self.spinBox_3.setMaximum(2100)
        self.spinBox_3.setObjectName("spinBox")
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setItalic(False)
        self.spinBox_3.setFont(font)
        self.spinBox_4 = QtWidgets.QSpinBox(self.frame_9)
        self.spinBox_4.setGeometry(QtCore.QRect(670, 10, 91, 22))
        self.spinBox_4.setStyleSheet("background-color: rgb(255, 255, 255);;\n"
                                          "color: rgb(0, 0, 0);")
        self.spinBox_4.setAlignment(QtCore.Qt.AlignCenter)
        self.spinBox_4.setMinimum(1)
        self.spinBox_4.setMaximum(52)
        self.spinBox_4.setObjectName("spinBox_4")
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setItalic(False)
        self.spinBox_4.setFont(font)
        self.calendarWidget_4.setSelectionMode(QtWidgets.QCalendarWidget.NoSelection)
        self.spinBox_4.setValue(QDate.currentDate().weekNumber()[0])
        self.spinBox_3.setValue(QDate.currentDate().weekNumber()[1])

        self.frame_50 = QtWidgets.QFrame(self.frame_3)
        self.frame_50.setGeometry(QtCore.QRect(10, 40, 781, 581))
        self.frame_50.setStyleSheet("color: rgb(255, 230, 207);")
        self.frame_50.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.frame_50.setFrameShadow(QtWidgets.QFrame.Plain)
        self.frame_50.setObjectName("frame_50")

        self.frame_70 = QtWidgets.QFrame(self.frame_7)
        self.frame_70.setGeometry(QtCore.QRect(10, 40, 781, 581))
        self.frame_70.setStyleSheet("color: rgb(255, 230, 207);")
        self.frame_70.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.frame_70.setFrameShadow(QtWidgets.QFrame.Plain)
        self.frame_70.setObjectName("frame_70")

        self.comboBox_3.setDisabled(True)
        self.comboBox_3.hide()

        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)


        self.comboBox_2.currentIndexChanged.connect(lambda: self.show_student_info())
        self.comboBox.currentTextChanged.connect(lambda: self.update_data())

        self.spinBox.valueChanged.connect(self.weekly_table_student)
        self.spinBox_2.valueChanged.connect(self.weekly_table_student)
        self.spinBox_3.valueChanged.connect(self.weekly_table_employee)
        self.spinBox_4.valueChanged.connect(self.weekly_table_employee)
        self.tabWidget.currentChanged.connect(self.tab_change_student)
        self.tabWidget_3.currentChanged.connect(self.tab_change_employee)
        font = QtGui.QFont()
        font.setPointSize(9)
        self.tableWidget_2.verticalHeader().setFont(font)
        self.tableWidget_2.horizontalHeader().setFont(font)
        font = QtGui.QFont()
        font.setPointSize(9)
        self.tableWidget_3.verticalHeader().setFont(font)
        self.tableWidget_3.horizontalHeader().setFont(font)

        self.dateEdit.dateChanged.connect(lambda: self.find_week_number())


        self.main_layout.addWidget(self.frame_400,0,0)
        self.layout_1 = QVBoxLayout()
        self.main_layout.addLayout(self.layout_1,0,1)
        self.layout_1.addWidget(self.frame_13)
        self.layout_1.addWidget(self.frame_2)
        self.layout_2 = QHBoxLayout()
        self.layout_1.addLayout(self.layout_2)
        self.spacer = QSpacerItem(0, 0, QSizePolicy.Expanding, QSizePolicy.Minimum)
        self.layout_2.addItem(self.spacer)
        self.layout_2.addWidget(self.pushButton_9)
        self.layout_2.addWidget(self.pushButton_8)
        self.layout_2.setContentsMargins(10,0,50,0)
        self.main_layout.setContentsMargins(20, 15, 20, 15)
        self.layout_3 = QVBoxLayout(self.frame_13)
        self.layout_4 = QHBoxLayout()
        self.layout_3.addLayout(self.layout_4)
        self.layout_4.addWidget(self.label_10)
        self.layout_4.addWidget(self.pushButton_14)
        self.layout_4.addWidget(self.pushButton_19)
        self.layout_4.addWidget(self.label_35)
        self.layout_4.addWidget(self.pushButton_20)
        self.layout_4.addWidget(self.dateEdit)
        self.layout_4.addWidget(self.pushButton_100)
        self.layout_3.addWidget(self.tableWidget)
        self.layout_4.setContentsMargins(10, 10, 10, 10)
        self.layout_1.setContentsMargins(0,0,30,10)
        self.layout_5 = QVBoxLayout(self.frame_2)
        self.layout_6 = QHBoxLayout()
        self.layout_7 = QVBoxLayout()
        self.layout_5.addLayout(self.layout_6)
        self.layout_5.addLayout(self.layout_7)
        self.layout_6.addWidget(self.label_8)
        self.layout_6.addWidget(self.comboBox)
        self.layout_6.addWidget(self.comboBox_2)
        self.layout_7.addWidget(self.frame_3)
        self.layout_7.addWidget(self.frame_7)
        self.layout_7.addWidget(self.frame_11)

        self.layout_8 = QVBoxLayout(self.frame_3)
        self.layout_10 = QHBoxLayout()
        self.layout_8.addLayout(self.layout_10)
        self.layout_10.addWidget(self.frame_301)
        self.layout_8.addWidget(self.frame_5,10)
        self.layout_8.addWidget(self.frame_4,10)
        self.layout_8.addWidget(self.frame_50,10)
        self.layout_9 = QHBoxLayout()
        self.layout_8.addLayout(self.layout_9)
        self.layout_9.addWidget(self.frame_302)
        self.layout_9.addItem(self.spacer)
        self.layout_9.addWidget(self.label_39)
        self.layout_9.addItem(self.spacer)
        self.layout_9.addWidget(self.label_40)

        self.layout_38 = QVBoxLayout(self.frame_7)
        self.layout_40 = QHBoxLayout()
        self.layout_38.addLayout(self.layout_40)
        self.layout_40.addWidget(self.frame_401)
        self.layout_38.addWidget(self.frame_8, 10)
        self.layout_38.addWidget(self.frame_9, 10)
        self.layout_38.addWidget(self.frame_70, 10)
        self.layout_39 = QHBoxLayout()
        self.layout_38.addLayout(self.layout_39)
        self.layout_39.addWidget(self.frame_402)
        self.layout_39.addItem(self.spacer)
        self.layout_39.addWidget(self.label_43)
        self.layout_39.addItem(self.spacer)
        self.layout_39.addWidget(self.label_89)

        self.layout_68 = QVBoxLayout(self.frame_11)
        self.layout_70 = QHBoxLayout()
        self.layout_68.addLayout(self.layout_70)
        self.layout_70.addWidget(self.label_32)
        self.layout_68.addWidget(self.frame_12, 10)
        self.layout_69 = QHBoxLayout()
        self.layout_68.addLayout(self.layout_69)
        self.layout_69.addWidget(self.frame_502)

        self.layout_11 = QVBoxLayout(self.frame_5)
        self.layout_12 = QHBoxLayout()
        self.layout_13 = QHBoxLayout()
        self.layout_14 = QVBoxLayout()
        self.layout_15 = QVBoxLayout()
        self.layout_16 = QVBoxLayout()
        self.layout_17 = QVBoxLayout()
        self.layout_12.setContentsMargins(0,10,0,10)
        self.layout_12.setSpacing(10)
        self.layout_13.setSpacing(10)
        self.layout_14.setSpacing(10)
        self.layout_15.setSpacing(10)
        self.layout_16.setSpacing(10)
        self.layout_17.setSpacing(10)
        self.layout_14.setAlignment(Qt.AlignHCenter)
        self.layout_16.setAlignment(Qt.AlignHCenter)
        self.layout_11.addLayout(self.layout_12)
        self.layout_11.addLayout(self.layout_13)
        self.layout_13.addLayout(self.layout_14)
        self.layout_13.addLayout(self.layout_15)
        self.layout_13.addLayout(self.layout_16)
        self.layout_13.addLayout(self.layout_17)
        self.layout_12.addWidget(self.label_25)
        self.layout_14.addWidget(self.label_45)
        self.layout_14.addWidget(self.label_47)
        self.layout_14.addWidget(self.label_44)
        self.layout_14.addWidget(self.label_48)
        self.layout_14.addWidget(self.label_50)
        self.layout_14.addWidget(self.label_51)
        self.layout_14.addWidget(self.label_46)
        self.layout_14.addWidget(self.label_61)
        self.layout_14.addWidget(self.label_64)
        self.layout_15.addWidget(self.label_52)
        self.layout_15.addWidget(self.label_53)
        self.layout_15.addWidget(self.label_70)
        self.layout_15.addWidget(self.label_71)
        self.layout_15.addWidget(self.label_72)
        self.layout_15.addWidget(self.label_73)
        self.layout_15.addWidget(self.label_74)
        self.layout_15.addWidget(self.label_84)
        self.layout_15.addWidget(self.label_82)
        self.layout_16.addWidget(self.label_55)
        self.layout_16.addWidget(self.label_56)
        self.layout_16.addWidget(self.label_57)
        self.layout_16.addWidget(self.label_58)
        self.layout_16.addWidget(self.label_59)
        self.layout_16.addWidget(self.label_60)
        self.layout_16.addWidget(self.label_49)
        self.layout_16.addWidget(self.label_68)
        self.layout_16.addWidget(self.label_65)
        self.layout_17.addWidget(self.label_54)
        self.layout_17.addWidget(self.label_75)
        self.layout_17.addWidget(self.label_76)
        self.layout_17.addWidget(self.comboBox_5)
        self.layout_17.addWidget(self.label_77)
        self.layout_17.addWidget(self.label_79)
        self.layout_17.addWidget(self.label_80)
        self.layout_17.addWidget(self.label_88)
        self.layout_17.addWidget(self.label_86)

        self.layout_41 = QVBoxLayout(self.frame_8)
        self.layout_42 = QHBoxLayout()
        self.layout_43 = QHBoxLayout()
        self.layout_44 = QVBoxLayout()
        self.layout_45 = QVBoxLayout()
        self.layout_46 = QVBoxLayout()
        self.layout_47 = QVBoxLayout()
        self.layout_48 = QHBoxLayout()
        self.layout_42.setContentsMargins(0, 10, 0, 10)
        self.layout_42.setSpacing(10)
        self.layout_43.setSpacing(10)
        self.layout_44.setSpacing(10)
        self.layout_45.setSpacing(10)
        self.layout_46.setSpacing(10)
        self.layout_47.setSpacing(10)
        self.layout_44.setAlignment(Qt.AlignHCenter)
        self.layout_46.setAlignment(Qt.AlignHCenter)
        self.layout_41.addLayout(self.layout_42)
        self.layout_41.addLayout(self.layout_43)
        self.layout_41.addLayout(self.layout_48)
        self.layout_43.addLayout(self.layout_44)
        self.layout_43.addLayout(self.layout_45)
        self.layout_43.addLayout(self.layout_46)
        self.layout_43.addLayout(self.layout_47)
        self.layout_42.addWidget(self.label_26)
        self.layout_44.addWidget(self.label_92)
        self.layout_44.addWidget(self.label_94)
        self.layout_44.addWidget(self.label_91)
        self.layout_44.addWidget(self.label_95)
        self.layout_44.addWidget(self.label_97)
        self.layout_44.addWidget(self.label_93)
        self.layout_45.addWidget(self.label_105)
        self.layout_45.addWidget(self.label_106)
        self.layout_45.addWidget(self.label_107)
        self.layout_45.addWidget(self.label_108)
        self.layout_45.addWidget(self.label_109)
        self.layout_45.addWidget(self.label_205)
        self.layout_46.addWidget(self.label_98)
        self.layout_46.addWidget(self.label_96)
        self.layout_46.addWidget(self.label_119)
        self.layout_46.addWidget(self.label_121)
        self.layout_46.addWidget(self.label_99)
        self.layout_46.addWidget(self.label_100)
        self.layout_46.addWidget(self.label_101)
        self.layout_47.addWidget(self.label_110)
        self.layout_47.addWidget(self.label_117)
        self.layout_47.addWidget(self.label_120)
        self.layout_47.addWidget(self.label_122)
        self.layout_47.addWidget(self.label_112)
        self.layout_47.addWidget(self.comboBox_9)
        self.layout_47.addWidget(self.label_114)
        self.layout_48.addWidget(self.label_130)

        self.layout_81 = QVBoxLayout(self.frame_12)
        self.layout_82 = QHBoxLayout()
        self.layout_83 = QHBoxLayout()
        self.layout_84 = QVBoxLayout()
        self.layout_85 = QVBoxLayout()
        self.layout_86 = QVBoxLayout()
        self.layout_87 = QVBoxLayout()
        self.layout_88 = QHBoxLayout()
        self.layout_82.setContentsMargins(0, 10, 0, 10)
        self.layout_82.setSpacing(10)
        self.layout_83.setSpacing(10)
        self.layout_84.setSpacing(10)
        self.layout_85.setSpacing(10)
        self.layout_86.setSpacing(10)
        self.layout_87.setSpacing(10)
        self.layout_84.setAlignment(Qt.AlignHCenter)
        self.layout_86.setAlignment(Qt.AlignHCenter)
        self.layout_81.addLayout(self.layout_82)
        self.layout_81.addLayout(self.layout_83)
        self.layout_81.addLayout(self.layout_88)
        self.layout_83.addLayout(self.layout_84)
        self.layout_83.addLayout(self.layout_85)
        self.layout_83.addLayout(self.layout_86)
        self.layout_83.addLayout(self.layout_87)
        self.layout_82.addWidget(self.label_29)
        self.layout_84.addWidget(self.label_171)
        self.layout_84.addWidget(self.label_173)
        self.layout_84.addWidget(self.label_170)
        self.layout_84.addWidget(self.label_174)
        self.layout_84.addWidget(self.label_175)
        self.layout_84.addWidget(self.label_172)
        self.layout_85.addWidget(self.label_179)
        self.layout_85.addWidget(self.label_180)
        self.layout_85.addWidget(self.label_181)
        self.layout_85.addWidget(self.label_182)
        self.layout_85.addWidget(self.label_183)
        self.layout_85.addWidget(self.label_206)
        self.layout_86.addWidget(self.label_189)
        self.layout_86.addWidget(self.label_186)
        self.layout_86.addWidget(self.label_190)
        self.layout_86.addWidget(self.label_192)
        self.layout_86.addWidget(self.label_176)
        self.layout_86.addWidget(self.label_177)
        self.layout_86.addWidget(self.label_178)
        self.layout_87.addWidget(self.label_188)
        self.layout_87.addWidget(self.label_187)
        self.layout_87.addWidget(self.label_191)
        self.layout_87.addWidget(self.label_193)
        self.layout_87.addWidget(self.label_184)
        self.layout_87.addWidget(self.comboBox_10)
        self.layout_87.addWidget(self.label_185)
        self.layout_88.addWidget(self.label_131)

        self.layout_21 = QVBoxLayout(self.frame_4)
        self.layout_22 = QHBoxLayout()
        self.layout_21.addLayout(self.layout_22)
        self.layout_22.addWidget(self.tabWidget)
        self.layout_22.addItem(self.spacer)
        self.layout_22.addWidget(self.spinBox)
        self.layout_22.addWidget(self.spinBox_2)
        self.layout_21.addWidget(self.calendarWidget_2)
        self.layout_21.addWidget(self.tableWidget_2)
        self.layout_21.addWidget(self.label_201)
        self.layout_23 = QVBoxLayout(self.frame_50)

        self.layout_51 = QVBoxLayout(self.frame_9)
        self.layout_52 = QHBoxLayout()
        self.layout_51.addLayout(self.layout_52)
        self.layout_52.addWidget(self.tabWidget_3)
        self.layout_52.addItem(self.spacer)
        self.layout_52.addWidget(self.spinBox_3)
        self.layout_52.addWidget(self.spinBox_4)
        self.layout_51.addWidget(self.calendarWidget_4)
        self.layout_51.addWidget(self.tableWidget_3)
        self.layout_51.addWidget(self.label_200)
        self.layout_53 = QVBoxLayout(self.frame_70)

        widgets = self.frame_5.findChildren(QWidget)
        for widget in widgets:
            # Do something with the widget
            widget.setStyleSheet("font-family: Arial; font-size: 12px; font-weight: bold; color: white; text-align: center;")

        widgets_2 = self.frame_8.findChildren(QWidget)
        for widget in widgets_2:
            # Do something with the widget
            widget.setStyleSheet(
                "font-family: Arial; font-size: 12px; font-weight: bold; color: white; text-align: center;")

        widgets_3 = self.frame_12.findChildren(QWidget)
        for widget in widgets_3:
            # Do something with the widget
            widget.setStyleSheet(
                "font-family: Arial; font-size: 12px; font-weight: bold; color: white; text-align: center;")

        self.find_week_number()
        self.load_txt()
        self.show_data()
        self.show_general_schedule()

        self.ready_to_close = False
        self.tableWidget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tableWidget.customContextMenuRequested.connect(self.show_context_menu)
        self.populate_table()
        self.step_fwd_weekly()
        self.step_back_weekly()

        def check_maximized():
            self.expand_general_schedule()
            if Dialog.isMaximized():
                Dialog.showNormal()
            else:
                Dialog.showMaximized()


        def close_dialog():
            self.log_out()
            if self.ready_to_close == True:
                Dialog.close()
            else:
                pass

        for row in range(self.tableWidget.rowCount()):
            for column in range(self.tableWidget.columnCount()):
                item = self.tableWidget.item(row, column)
                if item is None:
                    self.tableWidget.setItem(row, column, QtWidgets.QTableWidgetItem(''))

        for employee in data_objects.employees:
            if employee['title'] == 'Ogretmen' or  employee['title'] == 'Dil Konusmaci' or  employee['title'] == 'Psikolog':
                employee_full_list.append(data_objects.employees.index(employee))
            else:
                personel_full_list.append(data_objects.employees.index(employee))

        def check_debt(student):
            try:
                if len(student['student_attended']) > 0:
                    from datetime import datetime
                    debt_dates_month = []
                    debt_dates_year = []
                    calculated_debt = []
                    monthly_amount_debt = []
                    price_change_months = []
                    price_change_years = []
                    total_debt = []
                    last_debt = []
                    agreed_price = int(student['agreed_price'])
                    printer = ''
                    for check_date in student['student_attended']:
                        check_date_2 = check_date.split(' ')
                        given_date = check_date_2[-1]
                        # Parse the given date into a datetime object
                        parsed_date = datetime.strptime(given_date, "%d-%m-%Y")
                        # Get the current month and year as integers
                        current_month = datetime.now().month
                        current_year = datetime.now().year
                        reg_date = student['registration date']
                        parsed_date_2 = datetime.strptime(reg_date, "%d-%m-%Y")
                        if len(student['unpaid_debt']) > 0:
                            for check_added in student['unpaid_debt']:
                                if (str(parsed_date.year) + '-' + f"{parsed_date.month:02d}") in check_added:
                                    continue
                                else:
                                    if parsed_date.year < current_year or (parsed_date.year == current_year and parsed_date.month < current_month):
                                        if parsed_date_2.year < parsed_date.year or (parsed_date_2.year == parsed_date.year and parsed_date_2.month < parsed_date.month):
                                            debt_dates_year.append(parsed_date.year)
                                            debt_dates_month.append(parsed_date.month)

                        else:
                            debt_dates_year.append(parsed_date.year)
                            debt_dates_month.append(parsed_date.month)
                    if len(student['price_change']) > 0:
                        for check_date in student['price_change']:
                            check_date_2 = check_date.split('TL/')
                            given_date = check_date_2[-1]
                            # Parse the given date into a datetime object
                            parsed_date = datetime.strptime(given_date, "%d-%m-%Y")
                            price_change_years.append(parsed_date.year)
                            price_change_months.append(parsed_date.month)
                            monthly_amount_debt.append(int(check_date_2[0]))
            except ValueError:
                pass



                def calculate_debt(price_change_months, price_change_years, new_prices, debt_months, debt_years,
                                   agreed_price):
                    prices = {}
                    for i in range(len(price_change_months)):
                        prices[datetime(year=price_change_years[i], month=price_change_months[i], day=1)] = new_prices[
                            i]

                    sorted_debts = sorted(zip(debt_years, debt_months))
                    output = []
                    for debt in sorted_debts:
                        for date in sorted(prices.keys(), reverse=True):
                            if date <= datetime(year=debt[0], month=debt[1], day=1):
                                output.append(f"{prices[date]}TL/{debt[0]}-{debt[1]:02d}-{int(student['monthly_payment_date']):02d}")
                                break
                        else:
                            output.append(f"{agreed_price}TL/{debt[0]}-{debt[1]:02d}-{int(student['monthly_payment_date']):02d}")

                    return output

                def accumulate_prices(prices):
                    date_dict = defaultdict(int)
                    for price in prices:
                        price, date_str = price.split('TL/')
                        date = datetime.strptime(date_str, '%d-%m-%Y').date()
                        date_dict[date] += int(price)
                    return [f'{price}TL/{date}' for date, price in sorted(date_dict.items())]
                try:
                    if int(student['monthly_payment_date']) != 0:
                        total_debt = calculate_debt(price_change_months, price_change_years, monthly_amount_debt, debt_dates_month,
                                                              debt_dates_year, agreed_price)
                        last_debt = accumulate_prices(total_debt)
                    else:
                        last_debt = []
                        pass
                except ValueError:
                    pass
                debt2 = []
                debt_index = []
                for debt in last_debt:
                    debt_2 = debt.split('TL/')
                    debt2.append(debt_2[-1])
                for debt3 in debt2:
                    for unpaid in student['unpaid_debt']:
                        if debt3 in unpaid:
                            debt_index.append(debt2.index(debt3))
                        else:
                            pass
                for debt in last_debt:
                    if last_debt.index(debt) not in debt_index:
                        student['unpaid_debt'].append(debt)
        for stdnt in data_objects.students:
            if stdnt['agreed_price'] != "":
                check_debt(stdnt)
            else:
                pass
        with open("student_data.txt", "w", encoding="utf-8") as f:
            f.writelines(json.dumps(data_objects.students, default=str))
        with open('student_data.txt', 'r', encoding="utf-8") as f:
            data_objects.students = json.load(f)

        try:
            self.show_student_info()
        except IndexError:
            pass

        if data_objects.active_auth_level == 1:
            import datetime
            self.label_8.setText('Ust Yonetici Ekrani')
            self.pushButton_3.setEnabled(True)
            fixed_day = []
            student_with_debt = []
            for student in data_objects.students:
                if len(student['unpaid_debt']) > 0:
                    student_with_debt.append(student['name'] + ' ' + student['surname'])
                    fixed_day.append(student['monthly_payment_date'])
                else:
                    pass
            # Get the current date
            today = datetime.date.today()

            # Get the last day of the current month
            last_day = datetime.date(today.year, today.month, 1) + datetime.timedelta(days=32)
            last_day = last_day.replace(day=1) - datetime.timedelta(days=1)
            for day in fixed_day:
                try:
                    if today.day >= int(day) and len(student_with_debt) > 0:
                        msg_box = QMessageBox()
                        msg_box.setIcon(QMessageBox.Warning)
                        msg_box.setText('Odeme tarihi gecmis ve odemesi sisteme islenmemis ogrenciler var. Odemesi gecikmis ogrenci listesi: \n {}'.format('\n'.join(student_with_debt)))
                        msg_box.setWindowTitle('Uyarı')
                        msg_box.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
                        msg_box.setDefaultButton(QMessageBox.Cancel)
                        msg_box.exec()
                        break
                    else:
                        pass
                except ValueError:
                    pass

        elif data_objects.active_auth_level == 2:
            import datetime
            self.label_8.setText('Yonetici Ekrani')
            self.comboBox_10.setDisabled(True)
            self.comboBox_10.hide()
            self.comboBox_9.setDisabled(True)
            self.comboBox_9.hide()
            self.label_100.hide()
            self.label_177.hide()
            self.pushButton_3.setEnabled(True)
            fixed_day = []
            student_with_debt = []
            for student in data_objects.students:
                if len(student['unpaid_debt']) > 0:
                    student_with_debt.append(student['name'] + ' ' + student['surname'])
                    fixed_day.append(student['monthly_payment_date'])
                else:
                    pass
            # Get the current date
            today = datetime.date.today()

            # Get the last day of the current month
            last_day = datetime.date(today.year, today.month, 1) + datetime.timedelta(days=32)
            last_day = last_day.replace(day=1) - datetime.timedelta(days=1)
            for day in fixed_day:
                try:
                    if today.day >= int(day) and len(student_with_debt) > 0:
                        msg_box = QMessageBox()
                        msg_box.setIcon(QMessageBox.Warning)
                        msg_box.setText(
                            'Odeme tarihi gecmis ve odemesi sisteme islenmemis ogrenciler var. Odemesi gecikmis ogrenci listesi: \n {}'.format(
                                '\n'.join(student_with_debt)))
                        msg_box.setWindowTitle('Uyarı')
                        msg_box.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
                        msg_box.setDefaultButton(QMessageBox.Cancel)
                        msg_box.exec()
                        break
                    else:
                        pass
                except ValueError:
                    pass
        elif data_objects.active_auth_level == 3:
            self.label_8.setText('Sekreter Ekrani')
            self.comboBox.setCurrentText('Ogrenci Verileri')
            self.comboBox.setDisabled(True)
            self.pushButton_3.setDisabled(True)
            self.frame_11.hide()
            self.frame_7.hide()
        elif data_objects.active_auth_level == 4:
            self.label_8.setText('Ogretmen Ekrani')
            self.pushButton_2.setDisabled(True)
            self.pushButton_4.setDisabled(True)
            self.pushButton_7.setDisabled(True)
            self.pushButton_9.setDisabled(True)
            self.pushButton_3.setDisabled(True)
            self.frame_13.hide()
            self.frame_2.show()
            self.comboBox_2.setDisabled(True)
            self.comboBox_2.hide()
            self.comboBox_3.setEnabled(True)
            self.comboBox_3.show()
            self.frame_5.hide()
            self.comboBox_5.hide()
            self.label_58.hide()
            self.comboBox_9.hide()
            self.label_100.hide()
            self.label_80.hide()
            self.label_49.hide()
            for employee in data_objects.employees:
                if data_objects.user_logged_in['name'] == employee['name'] and data_objects.user_logged_in['surname'] == employee['surname']:
                    schedule_list.append(employee['teacher_schedule'])
                    schedule_list.append(employee['teacher_attended'])
                    schedule_list.append(employee['teacher_skipped'])
                    schedule_list.append(employee['schedule_changed'])
                    schedule_list.append(employee['schedule_cancelled'])

            for student in data_objects.students:
                if student['name'] + ' ' + student['surname'] in str(schedule_list):
                    restricted_student_list.append(student['name'] + ' ' + student['surname'])
                else:
                    pass

            self.comboBox.setCurrentIndex(1)
            self.comboBox.setCurrentIndex(0)

    def tab_change_student(self, index):
        if index == 0:
            self.tableWidget_2.hide()
            self.calendarWidget_2.show()
            self.spinBox.hide()
            self.spinBox_2.hide()
        else:
            self.calendarWidget_2.hide()
            self.weekly_table_student()
            self.tableWidget_2.show()
            self.spinBox.show()
            self.spinBox_2.show()


    def weekly_table_student(self):
        self.tableWidget_2.clear()
        days = ['PAZARTESI', 'SALI', 'CARSAMBA', 'PERSEMBE', 'CUMA', 'CUMARTESI', 'PAZAR']
        hours = ['9:30-10:10', '10:30-11:10', '11:30-12:10', '12:30-13:10', '13:10-14:00', '14:00-14:40',
                 '15:00-15:40', '16:00-16:40', '17:00-17:40']
        self.tableWidget_2.setHorizontalHeaderLabels(days)
        self.tableWidget_2.setVerticalHeaderLabels(hours)
        for student in data_objects.students:
            if self.comboBox_2.currentText() == student['name'] + ' ' + student['surname']:
                for day in student['student_schedule']:
                    try:
                        day_1 = day.split(' ')
                        day_2 = day_1[-1].split('-')
                        date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                        if date.weekNumber()[0] == self.spinBox_2.value() and date.weekNumber()[
                            1] == self.spinBox.value():
                            row = hours.index(day_1[4])
                            column = days.index(day_1[5])
                            item = QTableWidgetItem()
                            item.setText(day_1[0] + ' ' + day_1[1] + ' - ' + day_1[2] + ' ' + day_1[3])
                            item.setBackground(QColor('#FFD9A8'))
                            self.tableWidget_2.setItem(row, column, item)
                    except ValueError:
                        pass
                for day in student['student_attended']:
                    try:
                        day_1 = day.split(' ')
                        day_2 = day_1[-1].split('-')
                        date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                        if date.weekNumber()[0] == self.spinBox_2.value() and date.weekNumber()[
                            1] == self.spinBox.value():
                            row = hours.index(day_1[4])
                            column = days.index(day_1[5])
                            item = QTableWidgetItem()
                            item.setText(day_1[0] + ' ' + day_1[1] + ' - ' + day_1[2] + ' ' + day_1[3])
                            item.setBackground(QColor('#C5FFC5'))
                            self.tableWidget_2.setItem(row, column, item)
                    except ValueError:
                        pass
                for day in student['student_skipped']:
                    try:
                        day_1 = day.split(' ')
                        day_2 = day_1[-1].split('-')
                        date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                        if date.weekNumber()[0] == self.spinBox_2.value() and date.weekNumber()[
                            1] == self.spinBox.value():
                            row = hours.index(day_1[4])
                            column = days.index(day_1[5])
                            item = QTableWidgetItem()
                            item.setText(day_1[0] + ' ' + day_1[1] + ' - ' + day_1[2] + ' ' + day_1[3])
                            item.setBackground(QColor('#FF9292'))
                            self.tableWidget_2.setItem(row, column, item)
                    except ValueError:
                        pass
                for day in student['schedule_cancelled']:
                    try:
                        day_1 = day.split(' ')
                        day_2 = day_1[-1].split('-')
                        date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                        if date.weekNumber()[0] == self.spinBox_2.value() and date.weekNumber()[
                            1] == self.spinBox.value():
                            row = hours.index(day_1[7])
                            column = days.index(day_1[8])
                            item = QTableWidgetItem()
                            item.setText(
                                day_1[0] + ' ' + day_1[1] + ' ' + day_1[2] + ' - ' + day_1[3] + ' ' + day_1[4] + ' - ' +
                                day_1[5] + ' ' + day_1[6])
                            item.setBackground(QColor('#DFB2FE'))
                            self.tableWidget_2.setItem(row, column, item)
                    except ValueError:
                        pass

    def mark_dates_student(self):
        global lastly_selected_student
        for student in data_objects.students:
            if lastly_selected_student == student['name'] + ' ' + student['surname'] and len(data_objects.students) > 1:
                for day in student['student_schedule']:
                    day_1 = day.split(' ')
                    day_2 = day_1[-1].split('-')
                    date_format = QTextCharFormat()
                    date_format.setBackground(QColor('#FFFFFF'))
                    date_format.setToolTip('')
                    date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                    self.calendarWidget_2.setDateTextFormat(date, date_format)
                    self.calendarWidget_2.setSelectedDate(QDate())
                for day in student['student_attended']:
                    day_1 = day.split(' ')
                    day_2 = day_1[-1].split('-')
                    date_format = QTextCharFormat()
                    date_format.setBackground(QColor('#FFFFFF'))
                    date_format.setToolTip('')
                    date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                    self.calendarWidget_2.setDateTextFormat(date, date_format)
                    self.calendarWidget_2.setSelectedDate(QDate())
                for day in student['student_skipped']:
                    day_1 = day.split(' ')
                    day_2 = day_1[-1].split('-')
                    date_format = QTextCharFormat()
                    date_format.setBackground(QColor('#FFFFFF'))
                    date_format.setToolTip('')
                    date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                    self.calendarWidget_2.setDateTextFormat(date, date_format)
                    self.calendarWidget_2.setSelectedDate(QDate())
                for day in student['schedule_cancelled']:
                    day_1 = day.split(' ')
                    day_2 = day_1[-1].split('-')
                    date_format = QTextCharFormat()
                    date_format.setBackground(QColor('#FFFFFF'))
                    date_format.setToolTip('')
                    date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                    self.calendarWidget_2.setDateTextFormat(date, date_format)
                    self.calendarWidget_2.setSelectedDate(QDate())
            else:
                pass

        for student in data_objects.students:
            if self.comboBox_2.currentText() == student['name'] + ' ' + student['surname']:
                for day in student['student_schedule']:
                    day_1 = day.split(' ')
                    day_2 = day_1[-1].split('-')
                    date_format = QTextCharFormat()
                    date_format.setBackground(QColor('#FFD9A8'))
                    date_format.setToolTip(day)
                    date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                    self.calendarWidget_2.setDateTextFormat(date, date_format)
                for day in student['student_attended']:
                    day_1 = day.split(' ')
                    day_2 = day_1[-1].split('-')
                    date_format = QTextCharFormat()
                    date_format.setBackground(QColor('#C5FFC5'))
                    date_format.setToolTip(day)
                    date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                    self.calendarWidget_2.setDateTextFormat(date, date_format)
                for day in student['student_skipped']:
                    day_1 = day.split(' ')
                    day_2 = day_1[-1].split('-')
                    date_format = QTextCharFormat()
                    date_format.setBackground(QColor('#FF9292'))
                    date_format.setToolTip(day)
                    date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                    self.calendarWidget_2.setDateTextFormat(date, date_format)
                for day in student['schedule_cancelled']:
                    day_1 = day.split(' ')
                    day_2 = day_1[-1].split('-')
                    date_format = QTextCharFormat()
                    date_format.setBackground(QColor('#DFB2FE'))
                    date_format.setToolTip(day)
                    date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                    self.calendarWidget_2.setDateTextFormat(date, date_format)
        lastly_selected_student = self.comboBox_2.currentText()


    def tab_change_employee(self, index):
        if index == 0:
            self.tableWidget_3.hide()
            self.calendarWidget_4.show()
            self.spinBox_3.hide()
            self.spinBox_4.hide()
        else:
            self.calendarWidget_4.hide()
            self.weekly_table_employee()
            self.tableWidget_3.show()
            self.spinBox_3.show()
            self.spinBox_4.show()

    def weekly_table_employee(self):
        self.tableWidget_3.clear()
        days = ['PAZARTESI', 'SALI', 'CARSAMBA', 'PERSEMBE', 'CUMA', 'CUMARTESI', 'PAZAR']
        hours = ['9:30-10:10', '10:30-11:10', '11:30-12:10', '12:30-13:10', '13:10-14:00', '14:00-14:40',
                 '15:00-15:40', '16:00-16:40', '17:00-17:40']
        self.tableWidget_3.setHorizontalHeaderLabels(days)
        self.tableWidget_3.setVerticalHeaderLabels(hours)
        for employee in data_objects.employees:
            if self.comboBox_2.currentText() == employee['name'] + ' ' + employee['surname']:
                for day in employee['teacher_schedule']:
                    day_1 = day.split(' ')
                    day_2 = day_1[-1].split('-')
                    date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                    if date.weekNumber()[0] == self.spinBox_4.value() and date.weekNumber()[
                        1] == self.spinBox_3.value():
                        try:
                            my_string = day
                            day_of_week = my_string.split(' ')
                            list_start = my_string.index("[")
                            list_end = my_string.index("]") + 1
                            my_list_string = my_string[list_start:list_end]
                            my_list = eval(my_list_string)
                            group_class_list = my_list
                            for time in hours:
                                if time in day:
                                    row_no = hours.index(time)
                            if 'Grup Dersi' in day:
                                column_no = days.index(day_of_week[-2])
                                comboBox_11 = QtWidgets.QComboBox()
                                comboBox_11.addItem('Grup Dersi')
                                for student in group_class_list:
                                    comboBox_11.addItem(student)
                                self.tableWidget_3.setCellWidget(row_no, column_no, comboBox_11)
                                comboBox_11.setStyleSheet("background-color: #FFD9A8;")
                            elif 'Akademik Ders' in day:
                                column_no = days.index(day_of_week[-2])
                                label = QtWidgets.QLabel(group_class_list[0] + ' - Akademik Ders')
                                label.setAlignment(Qt.AlignCenter)
                                self.tableWidget_3.setCellWidget(row_no, column_no, label)
                                label.setStyleSheet("background-color: #FFD9A8;")
                                label.setWordWrap(True)
                            elif 'Attentioner Ders' in day:
                                column_no = days.index(day_of_week[-2])
                                label = QtWidgets.QLabel(group_class_list[0] + ' - Attentioner Ders')
                                label.setAlignment(Qt.AlignCenter)
                                self.tableWidget_3.setCellWidget(row_no, column_no, label)
                                label.setStyleSheet("background-color: #FFD9A8;")
                                label.setWordWrap(True)
                            else:
                                pass
                        except UnboundLocalError:
                            pass
                    else:
                        pass

                for day in employee['teacher_attended']:
                    day_1 = day.split(' ')
                    day_2 = day_1[-1].split('-')
                    date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                    if date.weekNumber()[0] == self.spinBox_4.value() and date.weekNumber()[
                        1] == self.spinBox_3.value():
                        try:
                            my_string = day
                            day_of_week = my_string.split(' ')
                            list_start = my_string.index("[")
                            list_end = my_string.index("]") + 1
                            my_list_string = my_string[list_start:list_end]
                            my_list = eval(my_list_string)
                            group_class_list = my_list
                            for time in hours:
                                if time in day:
                                    row_no = hours.index(time)
                            if 'Grup Dersi' in day:
                                column_no = days.index(day_of_week[-2])
                                comboBox_11 = QtWidgets.QComboBox()
                                comboBox_11.addItem('Grup Dersi')
                                for student in group_class_list:
                                    comboBox_11.addItem(student)
                                self.tableWidget_3.setCellWidget(row_no, column_no, comboBox_11)
                                comboBox_11.setStyleSheet("background-color: #C5FFC5;")
                            elif 'Akademik Ders' in day:
                                column_no = days.index(day_of_week[-2])
                                label = QtWidgets.QLabel(group_class_list[0] + ' - Akademik Ders')
                                label.setAlignment(Qt.AlignCenter)
                                self.tableWidget_3.setCellWidget(row_no, column_no, label)
                                label.setStyleSheet("background-color: #C5FFC5;")
                                label.setWordWrap(True)
                            elif 'Attentioner Ders' in day:
                                column_no = days.index(day_of_week[-2])
                                label = QtWidgets.QLabel(group_class_list[0] + ' - Attentioner Ders')
                                label.setAlignment(Qt.AlignCenter)
                                self.tableWidget_3.setCellWidget(row_no, column_no, label)
                                label.setStyleSheet("background-color: #C5FFC5;")
                                label.setWordWrap(True)
                            else:
                                pass
                        except UnboundLocalError:
                            pass
                    else:
                        pass

                for day in employee['teacher_skipped']:
                    day_1 = day.split(' ')
                    day_2 = day_1[-1].split('-')
                    date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                    if date.weekNumber()[0] == self.spinBox_4.value() and date.weekNumber()[
                        1] == self.spinBox_3.value():
                        try:
                            my_string = day
                            day_of_week = my_string.split(' ')
                            list_start = my_string.index("[")
                            list_end = my_string.index("]") + 1
                            my_list_string = my_string[list_start:list_end]
                            my_list = eval(my_list_string)
                            group_class_list = my_list
                            for time in hours:
                                if time in day:
                                    row_no = hours.index(time)
                            if 'Grup Dersi' in day:
                                column_no = days.index(day_of_week[-2])
                                comboBox_11 = QtWidgets.QComboBox()
                                comboBox_11.addItem('Grup Dersi')
                                for student in group_class_list:
                                    comboBox_11.addItem(student)
                                self.tableWidget_3.setCellWidget(row_no, column_no, comboBox_11)
                                comboBox_11.setStyleSheet("background-color: #FF9292;")
                            elif 'Akademik Ders' in day:
                                column_no = days.index(day_of_week[-2])
                                label = QtWidgets.QLabel(group_class_list[0] + ' - Akademik Ders')
                                label.setAlignment(Qt.AlignCenter)
                                self.tableWidget_3.setCellWidget(row_no, column_no, label)
                                label.setStyleSheet("background-color: #FF9292;")
                                label.setWordWrap(True)
                            elif 'Attentioner Ders' in day:
                                column_no = days.index(day_of_week[-2])
                                label = QtWidgets.QLabel(group_class_list[0] + ' - Attentioner Ders')
                                label.setAlignment(Qt.AlignCenter)
                                self.tableWidget_3.setCellWidget(row_no, column_no, label)
                                label.setStyleSheet("background-color: #FF9292;")
                                label.setWordWrap(True)
                            else:
                                pass
                        except UnboundLocalError:
                            pass
                    else:
                        pass

                for day in employee['schedule_cancelled']:
                    day_1 = day.split(' ')
                    day_2 = day_1[-1].split('-')
                    date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                    if date.weekNumber()[0] == self.spinBox_4.value() and date.weekNumber()[
                        1] == self.spinBox_3.value():
                        try:
                            my_string = day
                            day_of_week = my_string.split(' ')
                            list_start = my_string.index("[")
                            list_end = my_string.index("]") + 1
                            my_list_string = my_string[list_start:list_end]
                            my_list = eval(my_list_string)
                            group_class_list = my_list
                            for time in hours:
                                if time in day:
                                    row_no = hours.index(time)
                            if 'Grup Dersi' in day:
                                column_no = days.index(day_of_week[-2])
                                comboBox_11 = QtWidgets.QComboBox()
                                comboBox_11.addItem('Grup Dersi')
                                for student in group_class_list:
                                    comboBox_11.addItem(student)
                                self.tableWidget_3.setCellWidget(row_no, column_no, comboBox_11)
                                comboBox_11.setStyleSheet("background-color: #DFB2FE;")
                            elif 'Akademik Ders' in day:
                                column_no = days.index(day_of_week[-2])
                                label = QtWidgets.QLabel(group_class_list[0] + ' - Akademik Ders')
                                label.setAlignment(Qt.AlignCenter)
                                self.tableWidget_3.setCellWidget(row_no, column_no, label)
                                label.setStyleSheet("background-color: #DFB2FE;")
                                label.setWordWrap(True)
                            elif 'Attentioner Ders' in day:
                                column_no = days.index(day_of_week[-2])
                                label = QtWidgets.QLabel(group_class_list[0] + ' - Attentioner Ders')
                                label.setAlignment(Qt.AlignCenter)
                                self.tableWidget_3.setCellWidget(row_no, column_no, label)
                                label.setStyleSheet("background-color: #DFB2FE;")
                                label.setWordWrap(True)
                            else:
                                pass
                        except UnboundLocalError:
                            pass
                    else:
                        pass

    def mark_dates_employee(self):
        global lastly_selected_employee
        for employee in data_objects.employees:
            if lastly_selected_employee == employee['name'] + ' ' + employee['surname'] and len(
                    data_objects.employees) > 1:
                for day in employee['teacher_schedule']:
                    day_1 = day.split(' ')
                    day_2 = day_1[-1].split('-')
                    date_format = QTextCharFormat()
                    date_format.setBackground(QColor('#FFFFFF'))
                    date_format.setToolTip('')
                    date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                    self.calendarWidget_4.setDateTextFormat(date, date_format)
                    self.calendarWidget_4.setSelectedDate(QDate())
                for day in employee['teacher_attended']:
                    day_1 = day.split(' ')
                    day_2 = day_1[-1].split('-')
                    date_format = QTextCharFormat()
                    date_format.setBackground(QColor('#FFFFFF'))
                    date_format.setToolTip('')
                    date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                    self.calendarWidget_4.setDateTextFormat(date, date_format)
                    self.calendarWidget_4.setSelectedDate(QDate())
                for day in employee['teacher_skipped']:
                    day_1 = day.split(' ')
                    day_2 = day_1[-1].split('-')
                    date_format = QTextCharFormat()
                    date_format.setBackground(QColor('#FFFFFF'))
                    date_format.setToolTip('')
                    date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                    self.calendarWidget_4.setDateTextFormat(date, date_format)
                    self.calendarWidget_4.setSelectedDate(QDate())
                for day in employee['schedule_cancelled']:
                    day_1 = day.split(' ')
                    day_2 = day_1[-1].split('-')
                    date_format = QTextCharFormat()
                    date_format.setBackground(QColor('#FFFFFF'))
                    date_format.setToolTip('')
                    date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                    self.calendarWidget_4.setDateTextFormat(date, date_format)
                    self.calendarWidget_4.setSelectedDate(QDate())
            else:
                pass

        for employee in data_objects.employees:
            if self.comboBox_2.currentText() == employee['name'] + ' ' + employee['surname']:
                for day in employee['teacher_schedule']:
                    day_1 = day.split(' ')
                    day_2 = day_1[-1].split('-')
                    date_format = QTextCharFormat()
                    date_format.setBackground(QColor('#FFD9A8'))
                    date_format.setToolTip(day)
                    date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                    self.calendarWidget_4.setDateTextFormat(date, date_format)
                for day in employee['teacher_attended']:
                    day_1 = day.split(' ')
                    day_2 = day_1[-1].split('-')
                    date_format = QTextCharFormat()
                    date_format.setBackground(QColor('#C5FFC5'))
                    date_format.setToolTip(day)
                    date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                    self.calendarWidget_4.setDateTextFormat(date, date_format)
                for day in employee['teacher_skipped']:
                    day_1 = day.split(' ')
                    day_2 = day_1[-1].split('-')
                    date_format = QTextCharFormat()
                    date_format.setBackground(QColor('#FF9292'))
                    date_format.setToolTip(day)
                    date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                    self.calendarWidget_4.setDateTextFormat(date, date_format)
                for day in employee['schedule_cancelled']:
                    day_1 = day.split(' ')
                    day_2 = day_1[-1].split('-')
                    date_format = QTextCharFormat()
                    date_format.setBackground(QColor('#DFB2FE'))
                    date_format.setToolTip(day)
                    date = QDate(int(day_2[0]), int(day_2[1]), int(day_2[-1]))
                    self.calendarWidget_4.setDateTextFormat(date, date_format)
        lastly_selected_employee = self.comboBox_2.currentText()


    def set_student_statistics(self):
        group_planned = []
        academic_planned = []
        attentioner_planned = []
        group_attended = []
        academic_attended = []
        attentioner_attended = []
        group_skipped = []
        academic_skipped = []
        attentioner_skipped = []
        group_cancelled = []
        academic_cancelled = []
        attentioner_cancelled = []
        total_planned = 0
        total_attended = 0
        total_skipped = 0
        total_cancelled = 0
        total_changed = 0
        for student in data_objects.students:
            if self.comboBox_2.currentText() == student['name'] + ' ' + student['surname']:
                for schedule in student['student_schedule']:
                    if 'Grup' in schedule:
                        group_planned.append(schedule)
                    if 'Akademik' in schedule:
                        academic_planned.append(schedule)
                    if 'Attentioner' in schedule:
                        attentioner_planned.append(schedule)
                for schedule in student['student_attended']:
                    if 'Grup' in schedule:
                        group_attended.append(schedule)
                    if 'Akademik' in schedule:
                        academic_attended.append(schedule)
                    if 'Attentioner' in schedule:
                        attentioner_attended.append(schedule)
                for schedule in student['student_skipped']:
                    if 'Grup' in schedule:
                        group_skipped.append(schedule)
                    if 'Akademik' in schedule:
                        academic_skipped.append(schedule)
                    if 'Attentioner' in schedule:
                        attentioner_skipped.append(schedule)
                for schedule in student['schedule_cancelled']:
                    if 'Grup' in schedule:
                        group_cancelled.append(schedule)
                    if 'Akademik' in schedule:
                        academic_cancelled.append(schedule)
                    if 'Attentioner' in schedule:
                        attentioner_cancelled.append(schedule)
                total_attended = len(group_attended) + len(attentioner_attended) + len(academic_attended)
                total_skipped = len(group_skipped) + len(attentioner_skipped) + len(academic_skipped)
                total_cancelled = len(group_cancelled) + len(attentioner_cancelled) + len(academic_cancelled)
                total_changed = len(student['schedule_changed'])
                total_planned = len(student['student_schedule'])

        # Hide the y-axis labels
        # Set the x-axis label rotation angle and label position
        x = ["Toplam Planlanan Ders", "Toplam Katilim", "Toplam Devamsizlik", "Toplam Iptal Edilen",
             "Toplam Degistirilen",
             "Katilim Grup Dersi", "Katilim Akademik Ders", "Katilim Attentioner Ders", "Devamsizlik Grup Dersi",
             "Devamsizlik Akademik Ders", "Devamsizlik Attentioner Ders", "Iptal Grup Dersi", "Iptal Akademik Ders",
             "Iptal Attentioner Ders"]
        y = [total_planned, total_attended, total_skipped, total_cancelled, total_changed, len(group_attended),
             len(academic_attended), len(attentioner_attended), len(group_skipped), len(academic_skipped),
             len(attentioner_skipped), len(group_cancelled), len(academic_cancelled), len(attentioner_cancelled)]
        wrapped_labels = ['\n'.join(textwrap.wrap(label, width=20)) for label in x]
        colors = ['#FFBEBE', '#FFE2AD', '#DEFBBE', '#BEFBD9', '#D5F7FF', '#E1D5FF', '#F9D5FF', '#ECECEC', '#FFBEBE',
                  '#FFE2AD', '#DEFBBE', '#BEFBD9', '#D5F7FF', '#E1D5FF', '#F9D5FF', '#ECECEC']
        self.figure = plt.Figure()

        # Set the axes labels and add the FigureCanvas to the dialog
        self.canvas = FigureCanvas(self.figure)
        self.canvas.setParent(self.frame_50)
        self.canvas.setGeometry(QtCore.QRect(0, 50, 781, 450))
        self.barChart = self.figure.add_subplot(111)
        self.barChart.set_title("Ogrenci Istatistikleri", fontsize=8, pad=15)
        x_positions = np.arange(len(wrapped_labels))
        self.barChart.set_xticks(x_positions)
        self.barChart.set_xticklabels(wrapped_labels)
        bars = self.barChart.bar(x_positions, y, color=colors)


        self.barChart.spines['top'].set_visible(False)
        self.barChart.spines['right'].set_visible(False)
        self.barChart.spines['left'].set_visible(False)

        self.barChart.tick_params(axis='y', labelcolor='white')
        self.barChart.tick_params(axis='x', labelrotation=60, labelsize=7, pad=1)
        self.barChart.set_yticks([])
        self.barChart.set_yticklabels([])
        self.figure.patch.set_facecolor('#FCFBE1')
        self.barChart.set_facecolor('#FCFBE1')
        self.figure.subplots_adjust(bottom=0.3)
        # Add the value labels to the plot
        for bar in bars:
            height = bar.get_height()
            self.barChart.text(bar.get_x() + bar.get_width() / 2., height,
                               '%d' % int(height), ha='center', va='bottom')
        while self.layout_23.count():
            item = self.layout_23.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()
        self.layout_23.addWidget(self.canvas)
        self.canvas.show()

    def set_employee_statistics(self):
        group_planned = []
        academic_planned = []
        attentioner_planned = []
        group_attended = []
        academic_attended = []
        attentioner_attended = []
        group_skipped = []
        academic_skipped = []
        attentioner_skipped = []
        group_cancelled = []
        academic_cancelled = []
        attentioner_cancelled = []
        total_planned = 0
        total_attended = 0
        total_skipped = 0
        total_cancelled = 0
        total_changed = 0

        for employee in data_objects.employees:
            if self.comboBox_2.currentText() == employee['name'] + ' ' + employee['surname']:
                for schedule in employee['teacher_schedule']:
                    if 'Grup' in schedule:
                        group_planned.append(schedule)
                    if 'Akademik' in schedule:
                        academic_planned.append(schedule)
                    if 'Attentioner' in schedule:
                        attentioner_planned.append(schedule)
                for schedule in employee['teacher_attended']:
                    if 'Grup' in schedule:
                        group_attended.append(schedule)
                    if 'Akademik' in schedule:
                        academic_attended.append(schedule)
                    if 'Attentioner' in schedule:
                        attentioner_attended.append(schedule)
                for student in data_objects.students:
                    for schedule in student['student_skipped']:
                        if 'Grup' in schedule and (employee['name'] + ' ' + employee['surname']) in schedule:
                            group_skipped.append(schedule)
                        if 'Akademik' in schedule and (employee['name'] + ' ' + employee['surname']) in schedule:
                            academic_skipped.append(schedule)
                        if 'Attentioner' in schedule and (employee['name'] + ' ' + employee['surname']) in schedule:
                            attentioner_skipped.append(schedule)
                for schedule in employee['schedule_cancelled']:
                    if 'Grup' in schedule:
                        group_cancelled.append(schedule)
                    if 'Akademik' in schedule:
                        academic_cancelled.append(schedule)
                    if 'Attentioner' in schedule:
                        attentioner_cancelled.append(schedule)
                total_attended = len(group_attended) + len(attentioner_attended) + len(academic_attended)
                total_skipped = len(group_skipped) + len(attentioner_skipped) + len(academic_skipped)
                total_cancelled = len(group_cancelled) + len(attentioner_cancelled) + len(academic_cancelled)
                total_changed = len(employee['schedule_changed'])
                total_planned = len(employee['teacher_schedule'])

        # Hide the y-axis labels
        # Set the x-axis label rotation angle and label position
        x = ["Toplam Planlanan Ders", "Toplam Katilim", "Toplam Devamsizlik", "Toplam Iptal Edilen",
             "Toplam Degistirilen",
             "Katilim Grup Dersi", "Katilim Akademik Ders", "Katilim Attentioner Ders", "Devamsizlik Grup Dersi",
             "Devamsizlik Akademik Ders", "Devamsizlik Attentioner Ders", "Iptal Grup Dersi", "Iptal Akademik Ders",
             "Iptal Attentioner Ders"]
        y = [total_planned, total_attended, total_skipped, total_cancelled, total_changed, len(group_attended),
             len(academic_attended), len(attentioner_attended), len(group_skipped), len(academic_skipped),
             len(attentioner_skipped), len(group_cancelled), len(academic_cancelled), len(attentioner_cancelled)]
        wrapped_labels = ['\n'.join(textwrap.wrap(label, width=20)) for label in x]
        colors = ['#FFBEBE', '#FFE2AD', '#DEFBBE', '#BEFBD9', '#D5F7FF', '#E1D5FF', '#F9D5FF', '#ECECEC', '#FFBEBE',
                  '#FFE2AD', '#DEFBBE', '#BEFBD9', '#D5F7FF', '#E1D5FF', '#F9D5FF', '#ECECEC']
        self.figure = plt.Figure()

        # Set the axes labels and add the FigureCanvas to the dialog
        self.canvas = FigureCanvas(self.figure)
        self.canvas.setParent(self.frame_70)
        self.canvas.setGeometry(QtCore.QRect(0, 50, 781, 450))
        self.barChart = self.figure.add_subplot(111)
        self.barChart.set_title("Ogretmen Istatistikleri", fontsize=8, pad=15)
        x_positions = np.arange(len(wrapped_labels))
        self.barChart.set_xticks(x_positions)
        self.barChart.set_xticklabels(wrapped_labels)
        bars = self.barChart.bar(x_positions, y, color=colors)

        self.barChart.spines['top'].set_visible(False)
        self.barChart.spines['right'].set_visible(False)
        self.barChart.spines['left'].set_visible(False)

        self.barChart.tick_params(axis='y', labelcolor='white')
        self.barChart.tick_params(axis='x', labelrotation=60, labelsize=7, pad=1)
        self.barChart.set_yticks([])
        self.barChart.set_yticklabels([])
        self.figure.patch.set_facecolor('#FCFBE1')
        self.barChart.set_facecolor('#FCFBE1')
        self.figure.subplots_adjust(bottom=0.3)
        # Add the value labels to the plot
        for bar in bars:
            height = bar.get_height()
            self.barChart.text(bar.get_x() + bar.get_width() / 2., height,
                               '%d' % int(height), ha='center', va='bottom')
        while self.layout_53.count():
            item = self.layout_53.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()
        self.layout_53.addWidget(self.canvas)
        self.canvas.show()


    def save_file_dialog(self):
        global saving_option
        import tkinter as tk
        from tkinter import filedialog
        save_options.open_save_options()
        if saving_option == 1:
            root = tk.Tk()
            root.withdraw()  # hide the main window
            # open the file dialog and allow the user to select a directory
            file_path = filedialog.askdirectory(title="Select a directory to save the file")
            existing_file_path = 'ogrenci_verileri.xlsx'
            new_directory_path = file_path
            new_file_path = new_directory_path + "/student_data.xlsx"
            df = pd.read_csv("student_data.csv")
            writer = pd.ExcelWriter('ogrenci_verileri.xlsx')
            df.to_excel(writer, sheet_name='Sheet1', index=False)
            writer.save()
            filepath = os.path.join(new_directory_path, existing_file_path)
            # read the contents of the existing file
            new_file = os.path.join(new_directory_path, os.path.basename(existing_file_path))
            # copy the file to the new location
            shutil.copy2(existing_file_path, new_file)
            # write the contents to a new file in the selected directory
            # the selected directory will be stored in the `file_path` variable
                # Convert CSV to XLS using pandas
        if saving_option == 2:
            root = tk.Tk()
            root.withdraw()  # hide the main window
            # open the file dialog and allow the user to select a directory
            file_path = filedialog.askdirectory(title="Select a directory to save the file")
            existing_file_path = 'calisan_verileri.xlsx'
            new_directory_path = file_path
            new_file_path = new_directory_path + "/calisan_verileri.xlsx"
            df = pd.read_csv("employee_data.csv")
            writer = pd.ExcelWriter('calisan_verileri.xlsx')
            df.to_excel(writer, sheet_name='Sheet1', index=False)
            writer.save()
            filepath = os.path.join(new_directory_path, existing_file_path)
            # read the contents of the existing file
            new_file = os.path.join(new_directory_path, os.path.basename(existing_file_path))
            # copy the file to the new location
            shutil.copy2(existing_file_path, new_file)
            # write the contents to a new file in the selected directory
            # the selected directory will be stored in the `file_path` variable
            # Convert CSV to XLS using pandas
        else:
            pass

        if saving_option == 3:
            import openpyxl
            from openpyxl.utils.cell import get_column_letter
            from openpyxl.styles import Alignment

            # Create a new Excel workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # Get the data from the QTableWidget
            for row in range(self.tableWidget.rowCount()):
                # Skip rows 0, 10, 20, ..., 60
                if (row + 1) % 10 == 0:
                    continue

                # Write the row label
                label = self.tableWidget.verticalHeaderItem(row).text()
                cell = sheet.cell(row=row + 2, column=1)
                cell.value = label

                # Write the vertical headers to the first column
                if row == 0:
                    for i in range(self.tableWidget.rowCount()):
                        if (i + 1) % 11 == 0:
                            continue
                        label = self.tableWidget.verticalHeaderItem(i).text()
                        cell = sheet.cell(row=i + 2, column=1)
                        cell.value = label

                for column in range(self.tableWidget.columnCount()):
                    for row in range(self.tableWidget.rowCount()):
                        # Get the widget in the cell
                        item = self.tableWidget.cellWidget(row, column)

                        # Initialize the text variable with a default value
                        text = ""

                        # Get the text of the widget
                        if isinstance(item, QtWidgets.QLabel):
                            text = item.text()
                        elif isinstance(item, QtWidgets.QComboBox):
                            # Get the list of items in the QComboBox
                            items = [item.itemText(i) for i in range(item.count())]
                            # Combine the items into a comma-separated string
                            text = ", ".join(items)

                        # Write the text to the Excel sheet
                        cell = sheet.cell(row=row + 2, column=column + 2)
                        cell.value = text

                        # Write the column headers
                        if row == 0:
                            header = self.tableWidget.horizontalHeaderItem(column).text()
                            sheet.cell(row=1, column=column + 2).value = header
                        if row in range(0, 61, 10) and column == 0:
                            sheet.cell(row=row + 2, column=1).value = self.tableWidget.item(row, column).text()


            workbook.save('genel_cizelge.xlsx')
            root = tk.Tk()
            root.withdraw()  # hide the main window
            # open the file dialog and allow the user to select a directory
            file_path = filedialog.askdirectory(title="Select a directory to save the file")
            existing_file_path = 'genel_cizelge.xlsx'
            new_directory_path = file_path
            filepath = os.path.join(new_directory_path, existing_file_path)
            # read the contents of the existing file
            new_file = os.path.join(new_directory_path, os.path.basename(existing_file_path))
            # copy the file to the new location
            shutil.copy2(existing_file_path, new_file)



    def load_txt(self):
        try:
            if data_objects.active_auth_level == 3:
                self.comboBox.addItems(['Ogrenci Verileri'])
            elif data_objects.active_auth_level == 4:
                self.comboBox.addItems(['Ogrenci Verileri', 'Ogretmen Verileri'])
            else:
                self.comboBox.addItems(['Ogrenci Verileri', 'Ogretmen Verileri', 'Personel Verileri'])

            with open('employee_data.txt', 'r', encoding="utf-8") as f:
                data_objects.employees = json.load(f)

            with open('student_data.txt', 'r', encoding="utf-8") as f:
                data_objects.students = json.load(f)

            for student in data_objects.students:
                self.comboBox_2.addItem((student['name']) + ' ' + (student['surname']))
                self.comboBox_2.setItemData(0,(student['name']) + ' ' + (student['surname']).replace("\n", "\\n"), role=Qt.DisplayRole)
                student_list.append((student['name']) + ' ' + (student['surname']))
                self.comboBox_2.model().sort(0)
            self.comboBox_3.addItems(restricted_student_list)
        except json.decoder.JSONDecodeError:
            pass

    def update_data(self):
        try:
            with open('employee_data.txt', 'r', encoding="utf-8") as f:
                data_objects.employees = json.load(f)

            with open('student_data.txt', 'r', encoding="utf-8") as f:
                data_objects.students = json.load(f)

            self.combobox_top()

            if self.comboBox.currentText() == 'Ogrenci Verileri':
                self.comboBox_2.currentIndexChanged.connect(lambda: self.show_student_info())
                self.comboBox_2.clear()
                self.comboBox_3.currentIndexChanged.connect(lambda: self.show_student_info())
                self.comboBox_3.clear()
                for student in data_objects.students:
                    self.comboBox_2.addItem((student['name']) + ' ' + (student['surname']))
                self.comboBox_3.addItems(restricted_student_list)
            elif self.comboBox.currentText() == 'Ogretmen Verileri':
                self.comboBox_2.currentIndexChanged.connect(lambda: self.show_teacher_info())
                self.comboBox_2.clear()
                self.comboBox_3.currentIndexChanged.connect(lambda: self.show_teacher_info())
                self.comboBox_3.clear()
                for employee in data_objects.employees:
                    if employee['title'] == 'Ogretmen' or  employee['title'] == 'Dil Konusmaci' or  employee['title'] == 'Psikolog':
                        self.comboBox_2.addItem((employee['name']) + ' ' + (employee['surname']))

                    else:
                        pass
                self.comboBox_3.addItem(data_objects.user_logged_in['name'] + ' ' + data_objects.user_logged_in['surname'])
            else:
                self.comboBox_2.currentIndexChanged.connect(lambda: self.show_personel_info())
                self.comboBox_2.clear()
                for employee in data_objects.employees:
                    if employee['title'] == 'Ogretmen' or  employee['title'] == 'Dil Konusmaci' or  employee['title'] == 'Psikolog':
                        pass
                    else:
                        self.comboBox_2.addItem((employee['name']) + ' ' + (employee['surname']))
            self.comboBox_2.model().sort(0)
        except json.decoder.JSONDecodeError:
            pass

    def show_student_info(self):
        global data_count_student
        self.frame_4.hide()
        self.frame_5.show()
        self.frame_50.hide()
        self.label_30.setText("  Ogrenci Bilgileri Genel Ozet")
        self.tabWidget.setCurrentIndex(0)
        data_count_student = 2
        selected = ''
        try:
            with open('employee_data.txt', 'r', encoding="utf-8") as f:
                data_objects.employees = json.load(f)
            with open('student_data.txt', 'r', encoding="utf-8") as f:
                data_objects.students = json.load(f)
            if data_objects.active_auth_level == 4:
                selected = self.comboBox_3.currentText()
            else:
                selected = self.comboBox_2.currentText()
            selected_list = selected.split(' ')
            i = 0
            for student in data_objects.students:
                if selected_list[0] in student['name'] and selected_list[-1] in student['surname']:
                    i = data_objects.students.index(student)
                    break
                else:
                    i = 0
            self.label_25.setPixmap(QtGui.QPixmap("default-user-image.png"))
            self.label_52.setText(data_objects.students[i]['name'])
            self.label_53.setText(data_objects.students[i]['surname'])
            self.label_70.setText(data_objects.students[i]['date_of_birth'])
            self.label_71.setText(data_objects.students[i]['identity_no'])
            self.label_72.setText(data_objects.students[i]['school'])
            self.label_73.setText(data_objects.students[i]['class'])
            self.label_74.setText(str(data_objects.students[i]['former_support']))
            self.label_54.setText(data_objects.students[i]['registeration_type'])
            self.label_75.setText(data_objects.students[i]['module'])
            self.label_76.setText(data_objects.students[i]['transportation_service'])
            self.comboBox_5.clear()
            price_change = data_objects.students[i]['price_change'][::-1]
            for price in price_change:
                self.comboBox_5.addItem(price)
            self.comboBox_5.addItem('                '+data_objects.students[i]['agreed_price'] + 'TL' + ' / ' + data_objects.students[i]['registration date'])
            self.label_77.setText(data_objects.students[i]['registration date'])
            self.label_79.setText(data_objects.students[i]['notes_problems'])
            self.label_80.setText(data_objects.students[i]['address'])
            self.label_84.setText(data_objects.students[i]['mother_name'])
            self.label_82.setText(data_objects.students[i]['mother_phone'])
            self.label_88.setText(data_objects.students[i]['father_name'])
            self.label_86.setText(data_objects.students[i]['father_phone'])
            self.label_39.setText("Toplam Katilim: {} Ders".format(len(data_objects.students[i]['student_attended'])))
            self.label_40.setText("Devamsizlik: {} Ders".format(len(data_objects.students[i]['student_skipped'])))
            if data_objects.students[i]['status'][0] == 'A':
                self.label_41.setText("Durum: {}".format(data_objects.students[i]['status']))
                self.label_38.setPixmap(QtGui.QPixmap("yes.png"))
                self.label_38.raise_()
            else:
                stat = (data_objects.students[i]['status'] + ' ' + '(' + data_objects.students[i]['student_left'] +')')
                self.label_41.setText("Durum: {}".format(stat))
                self.label_38.setPixmap(QtGui.QPixmap("delete.png"))
                self.label_38.raise_()

        except IndexError:
            pass

    def show_teacher_info(self):
        global data_count_teacher
        self.frame_9.hide()
        self.frame_8.show()
        self.label_118.show()
        self.frame_70.hide()
        self.label_31.setText(" Ogretmen Bilgileri Genel Ozet")
        self.tabWidget_3.setCurrentIndex(0)
        data_count_teacher = 2
        selected_teacher = ''
        try:
            with open('employee_data.txt', 'r', encoding="utf-8") as f:
                data_objects.employees = json.load(f)

            with open('student_data.txt', 'r', encoding="utf-8") as f:
                data_objects.students = json.load(f)
            if data_objects.active_auth_level == 4:
                selected_teacher = self.comboBox_3.currentText()
            else:
                selected_teacher = self.comboBox_2.currentText()
            i = 0
            for employee in data_objects.employees:
                if employee['name'] in selected_teacher and employee['surname'] in selected_teacher:
                    i = data_objects.employees.index(employee)
            self.label_26.setPixmap(QtGui.QPixmap("default-user-image.png"))
            self.label_105.setText(data_objects.employees[i]['name'])
            self.label_106.setText(data_objects.employees[i]['surname'])
            self.label_107.setText(data_objects.employees[i]['date_of_birth'])
            self.label_108.setText(data_objects.employees[i]['identity_no'])
            self.label_109.setText(data_objects.employees[i]['graduation_school'])
            self.label_110.setText(data_objects.employees[i]['phone'])
            self.label_117.setText(data_objects.employees[i]['address'])
            self.label_122.setText(data_objects.employees[i]['email'])
            self.label_205.setText(str(data_objects.employees[i]['experience']))
            self.comboBox_9.clear()
            salary_change = data_objects.employees[i]['salary_change'][::-1]
            for salary in salary_change:
                self.comboBox_9.addItem(salary)
            self.comboBox_9.addItem(data_objects.employees[i]['agreed_salary'] + 'TL' + ' / ' + data_objects.employees[i]['registration_date'])
            self.label_112.setText(data_objects.employees[i]['type_employment'])
            self.label_120.setText(data_objects.employees[i]['graduation_date'])
            self.label_80.setText(data_objects.employees[i]['address'])
            self.label_114.setText(data_objects.employees[i]['registration_date'])
            self.label_130.setText("Calisma Gunler: {}".format(str(data_objects.employees[i]['working_days'])))
            self.label_43.setText("Toplam Katilim: {} Ders".format(len(data_objects.employees[i]['teacher_attended'])))
            self.label_89.setText("Iptal Olan Ders: {}".format(len(data_objects.employees[i]['schedule_cancelled'])))
            if data_objects.employees[i]['status'][0] == 'A':
                self.label_90.setText("Durum: {}".format(data_objects.employees[i]['status']))
                self.label_42.setPixmap(QtGui.QPixmap("yes.png"))
                self.label_42.raise_()
            else:
                self.label_90.setText("Durum: {}".format(data_objects.employees[i]['status']))
                self.label_42.setPixmap(QtGui.QPixmap("delete.png"))
                self.label_42.raise_()

        except IndexError:
            pass

    def show_personel_info(self):
        try:
            with open('employee_data.txt', 'r', encoding="utf-8") as f:
                data_objects.employees = json.load(f)
            with open('student_data.txt', 'r', encoding="utf-8") as f:
                data_objects.students = json.load(f)
            if data_objects.active_auth_level == 4:
                selected_teacher = self.comboBox_3.currentText()
            else:
                selected_teacher = self.comboBox_2.currentText()
            i = 0
            for employee in data_objects.employees:
                if employee['name'] in selected_teacher and employee['surname'] in selected_teacher:
                    i = data_objects.employees.index(employee)
            self.label_29.setPixmap(QtGui.QPixmap("default-user-image.png"))
            self.label_179.setText(data_objects.employees[i]['name'])
            self.label_180.setText(data_objects.employees[i]['surname'])
            self.label_181.setText(data_objects.employees[i]['date_of_birth'])
            self.label_182.setText(data_objects.employees[i]['identity_no'])
            self.label_183.setText(data_objects.employees[i]['graduation_school'])
            self.label_188.setText(data_objects.employees[i]['phone'])
            self.label_187.setText(data_objects.employees[i]['address'])
            self.label_193.setText(data_objects.employees[i]['email'])
            self.comboBox_10.clear()
            self.comboBox_10.addItem('                '+data_objects.employees[i]['agreed_salary'] + 'TL' + ' / ' + data_objects.employees[i]['registration_date'])
            self.label_184.setText(data_objects.employees[i]['type_employment'])
            self.label_191.setText(data_objects.employees[i]['graduation_date'])
            self.label_185.setText(data_objects.employees[i]['registration_date'])
            self.label_206.setText(str(data_objects.employees[i]['experience']))
            self.label_131.setText("Calisma Gunler: {}".format(str(data_objects.employees[i]['working_days'])))
            if data_objects.employees[i]['status'][0] == 'A':
                self.label_169.setText("Durum: {}".format(data_objects.employees[i]['status']))
                self.label_113.setPixmap(QtGui.QPixmap("yes.png"))
                self.label_113.raise_()
            else:
                self.label_169.setText("Durum: {}".format(data_objects.employees[i]['status']))
                self.label_113.setPixmap(QtGui.QPixmap("delete.png"))
                self.label_113.raise_()
        except IndexError:
            pass

    def expand_general_schedule(self):
        global row_height, row_extender, section_size
        try:
            section_size = int(max(row_height))
            if row_extender == 0:
                self.tableWidget.verticalHeader().setMinimumSectionSize(35 + section_size * 15)
                row_extender = 1
            else:
                self.tableWidget.verticalHeader().setMinimumSectionSize(50)
                row_extender = 0
        except ValueError:
            pass

    def get_maximum_content_height(self, row):
        max_height = 0
        for col in range(self.tableWidget.columnCount()):
            item = self.tableWidget.item(row, col)
            if item is not None:
                cell_height = self.tableWidget.rowHeight(row)
                content_height = item.sizeHint().height()
                max_height = max(max_height, content_height, cell_height)
        return max_height

    def load_general_schedule(self):
        global group_class_list, date_list, row_height
        self.tableWidget.verticalHeader().setMinimumSectionSize(50)
        days = ['PAZARTESI 20', 'SALI 20', 'CARSAMBA 20', 'PERSEMBE 20', 'CUMA 20', 'CUMARTESI 20', 'PAZAR 20']
        v_header = [' ', '9:30-10:10', '10:30-11:10', '11:30-12:10', '12:30-13:10', '13:10-14:00', '14:00-14:40',
                    '15:00-15:40', '16:00-16:40', '17:00-17:40']
        header_list = []
        for i in range(self.tableWidget.columnCount()):
            header_item = self.tableWidget.horizontalHeaderItem(i)  # retrieve header item for column i
            if header_item is not None:
                header_list.append(header_item.text())  # add header item text to the list
        for employee in data_objects.employees:
            for schedule in employee['teacher_schedule']:
                if any(elem in schedule for elem in date_list):
                    my_string = schedule
                    list_start = my_string.index("[")
                    list_end = my_string.index("]") + 1
                    my_list_string = my_string[list_start:list_end]
                    my_list = eval(my_list_string)
                    group_class_list = my_list
                    row_height.append(len(group_class_list))
                    for day in days:
                        if day in schedule:
                            session = days.index(day)
                    for time in v_header:
                        if time in schedule:
                            row_no = v_header.index(time) + session * 10

                    if 'Grup Dersi' in schedule:
                        column_no = header_list.index((employee['name']+' '+employee['surname']))
                        item = QTableWidgetItem('GRUP DERSI:\n{}'.format('\n'.join(group_class_list)))
                        item.setData(Qt.BackgroundRole, QColor("#FFD9A8"))
                        item.setTextAlignment(Qt.AlignCenter)
                        self.tableWidget.setItem(row_no, column_no, item)

                    elif 'Akademik Ders' in schedule:
                        column_no = header_list.index((employee['name'] + ' ' + employee['surname']))
                        item = QTableWidgetItem(f"{group_class_list[0]} - AKADEMIK DERS")
                        item.setData(Qt.BackgroundRole, QColor("#FFD9A8"))
                        item.setTextAlignment(Qt.AlignCenter)  # AlignCenter
                        self.tableWidget.setItem(row_no, column_no, item)

                    elif 'Attentioner Ders' in schedule:
                        column_no = header_list.index((employee['name'] + ' ' + employee['surname']))
                        item = QTableWidgetItem(f"{group_class_list[0]} - ATTENTIONER DERS")
                        item.setData(Qt.BackgroundRole, QColor("#FFD9A8"))
                        item.setTextAlignment(Qt.AlignCenter)  # AlignCenter
                        self.tableWidget.setItem(row_no, column_no, item)
                    else:
                        pass
                else:
                    pass

            for schedule in employee['schedule_cancelled']:
                if any(elem in schedule for elem in date_list):
                    my_string = schedule
                    list_start = my_string.index("[")
                    list_end = my_string.index("]") + 1
                    my_list_string = my_string[list_start:list_end]
                    my_list = eval(my_list_string)
                    group_class_list = my_list
                    for day in days:
                        if day in schedule:
                            session = days.index(day)
                    for time in v_header:
                        if time in schedule:
                            row_no = v_header.index(time) + session * 10

                    if 'Grup Dersi' in schedule:
                        column_no = header_list.index((employee['name'] + ' ' + employee['surname']))
                        item = QTableWidgetItem('GRUP DERSI: \n {}'.format('\n'.join(group_class_list)))
                        item.setData(Qt.BackgroundRole, QColor("#FF9292"))
                        item.setTextAlignment(Qt.AlignCenter)
                        self.tableWidget.setItem(row_no, column_no, item)
                    elif 'Akademik Ders' in schedule:
                        column_no = header_list.index((employee['name'] + ' ' + employee['surname']))
                        item = QTableWidgetItem(f"{group_class_list[0]} - AKADEMIK DERS")
                        item.setData(Qt.BackgroundRole, QColor("#FF9292"))
                        item.setTextAlignment(Qt.AlignCenter)  # AlignCenter
                        self.tableWidget.setItem(row_no, column_no, item)
                    elif 'Attentioner Ders' in schedule:
                        column_no = header_list.index((employee['name'] + ' ' + employee['surname']))
                        item = QTableWidgetItem(f"{group_class_list[0]} - ATTENTIONER DERS")
                        item.setData(Qt.BackgroundRole, QColor("#FF9292"))
                        item.setTextAlignment(Qt.AlignCenter)  # AlignCenter
                        self.tableWidget.setItem(row_no, column_no, item)

                    else:
                        pass
                else:
                    pass

            for schedule in employee['teacher_attended']:
                if any(elem in schedule for elem in date_list):
                    my_string = schedule
                    list_start = my_string.index("[")
                    list_end = my_string.index("]") + 1
                    my_list_string = my_string[list_start:list_end]
                    my_list = eval(my_list_string)
                    group_class_list = my_list
                    for day in days:
                        if day in schedule:
                            session = days.index(day)
                    for time in v_header:
                        if time in schedule:
                            row_no = v_header.index(time) + session * 10

                    if 'Grup Dersi' in schedule:
                        column_no = header_list.index((employee['name'] + ' ' + employee['surname']))
                        item = QTableWidgetItem('GRUP DERSI: \n {}'.format('\n'.join(group_class_list)))
                        item.setData(Qt.BackgroundRole, QColor("#C5FFC5"))
                        item.setTextAlignment(Qt.AlignCenter)
                        self.tableWidget.setItem(row_no, column_no, item)
                    elif 'Akademik Ders' in schedule:
                        column_no = header_list.index((employee['name'] + ' ' + employee['surname']))
                        item = QTableWidgetItem(f"{group_class_list[0]} - AKADEMIK DERS")
                        item.setData(Qt.BackgroundRole, QColor("#C5FFC5"))
                        item.setTextAlignment(Qt.AlignCenter)  # AlignCenter
                        self.tableWidget.setItem(row_no, column_no, item)
                    elif 'Attentioner Ders' in schedule:
                        column_no = header_list.index((employee['name'] + ' ' + employee['surname']))
                        item = QTableWidgetItem(f"{group_class_list[0]} - ATTENTIONER DERS")
                        item.setData(Qt.BackgroundRole, QColor("#C5FFC5"))
                        item.setTextAlignment(Qt.AlignCenter)  # AlignCenter
                        self.tableWidget.setItem(row_no, column_no, item)

                    else:
                        pass
                else:
                    pass
        self.tableWidget.customContextMenuRequested.connect(self.show_context_menu)

    def populate_table(self):
        for row in range(self.tableWidget.rowCount()):
            for col in range(self.tableWidget.columnCount()):
                item = QTableWidgetItem()
                self.tableWidget.setItem(row, col, item)


    def show_context_menu(self, pos):
        cell = self.tableWidget.itemAt(pos)
        if cell is not None:
            menu = QMenu(self.tableWidget)
            menu.setStyleSheet("QMenu { background-color: #F0F0F0; color: #000000; }"
                               "QMenu::item { background-color: #F0F0F0; color: #000000; }"
                               "QMenu::item:selected { background-color: #3366FF; color: #FFFFFF; }")


            action1 = QAction("Akademik Ders Ekle", self.tableWidget)
            action1.triggered.connect(lambda: self.handle_context_menu_action("Akademik Ders Ekle", cell))
            menu.addAction(action1)

            action2 = QAction("Attentioner Ders Ekle", self.tableWidget)
            action2.triggered.connect(lambda: self.handle_context_menu_action("Attentioner Ders Ekle", cell))
            menu.addAction(action2)

            action3 = QAction("Grup Dersi Ekle", self.tableWidget)
            action3.triggered.connect(lambda: self.handle_context_menu_action("Grup Dersi Ekle", cell))
            menu.addAction(action3)

            action4 = QAction("Dersi Iptal Et", self.tableWidget)
            action4.triggered.connect(lambda: self.handle_context_menu_action("Dersi Iptal Et", cell))
            menu.addAction(action4)

            action5 = QAction("Yoklama Bilgisi Ekle", self.tableWidget)
            action5.triggered.connect(lambda: self.handle_context_menu_action("Yoklama Bilgisi Ekle", cell))
            menu.addAction(action5)

            action6 = QAction("Aktar (1 Hafta)", self.tableWidget)
            action6.triggered.connect(lambda: self.handle_context_menu_action("Aktar (1 Hafta)", cell))
            menu.addAction(action6)

            action7 = QAction("Dersi Sil", self.tableWidget)
            action7.triggered.connect(lambda: self.handle_context_menu_action("Dersi Sil", cell))
            menu.addAction(action7)

            cell_color = cell.background().color().name()

            if cell_color == "#ffd9a8":
                action1.setDisabled(True)
                action2.setDisabled(True)
                action3.setDisabled(True)
            elif cell_color == "#c5ffc5":
                action1.setDisabled(True)
                action2.setDisabled(True)
                action3.setDisabled(True)
                action4.setDisabled(True)
                action5.setDisabled(True)
            elif cell_color == "#ff9292":
                action1.setDisabled(True)
                action2.setDisabled(True)
                action3.setDisabled(True)
                action4.setDisabled(True)
                action5.setDisabled(True)
                action6.setDisabled(True)
            else:
                action1.setEnabled(True)
                action2.setEnabled(True)
                action3.setEnabled(True)
                action4.setDisabled(True)
                action5.setDisabled(True)
                action6.setDisabled(True)
                action7.setDisabled(True)
            


            menu.exec_(self.tableWidget.viewport().mapToGlobal(pos))

    def handle_context_menu_action(self, option, cell):
        global group_class_list, date_list
        group_class_list = []
        if cell.row() in range(0, 10):
            item = self.tableWidget.item(0, 0).text()
        elif cell.row() in range(10, 20):
            item = self.tableWidget.item(10, 0).text()
        elif cell.row() in range(20, 30):
            item = self.tableWidget.item(20, 0).text()
        elif cell.row() in range(30, 40):
            item = self.tableWidget.item(30, 0).text()
        elif cell.row() in range(40, 50):
            item = self.tableWidget.item(40, 0).text()
        elif cell.row() in range(50, 60):
            item = self.tableWidget.item(50, 0).text()
        elif cell.row() in range(60, 70):
            item = self.tableWidget.item(60, 0).text()
        else:
            pass

        if option == "Akademik Ders Ekle":
            class_options.class_type = 1
            class_options.open_class_options()
            if class_options.if_cancelled == 0:
                for student in data_objects.students:
                    if (student['name'] + ' ' + student['surname']) in group_class_list:
                        student['student_schedule'].append(
                            'Akademik Ders' + ' ' + self.tableWidget.horizontalHeaderItem(cell.column()).text() +
                            ' ' + self.tableWidget.verticalHeaderItem(cell.row()).text() + ' ' + item)
                for employee in data_objects.employees:
                    if (employee['name'] + ' ' + employee['surname']) == self.tableWidget.horizontalHeaderItem(
                            cell.column()).text():
                        employee['teacher_schedule'].append(('Akademik Ders' + ' ' + str(group_class_list) + ' ' +
                                                             self.tableWidget.verticalHeaderItem(cell.row()).text() + ' ' + item))
            else:
                pass
        elif option == "Attentioner Ders Ekle":
            class_options.class_type = 2
            class_options.open_class_options()
            if class_options.if_cancelled == 0:
                for student in data_objects.students:
                    if (student['name'] + ' ' + student['surname']) in group_class_list:
                        student['student_schedule'].append(
                            'Attentioner Ders' + ' ' + self.tableWidget.horizontalHeaderItem(cell.column()).text() +
                            ' ' + self.tableWidget.verticalHeaderItem(cell.row()).text() + ' ' + item)
                for employee in data_objects.employees:
                    if (employee['name'] + ' ' + employee['surname']) == self.tableWidget.horizontalHeaderItem(
                            cell.column()).text():
                        employee['teacher_schedule'].append(('Attentioner Ders' + ' ' + str(group_class_list) + ' ' +
                                                             self.tableWidget.verticalHeaderItem(cell.row()).text() + ' ' + item))
            else:
                pass

        elif option == "Grup Dersi Ekle":
            class_options.class_type = 3
            class_options.open_class_options()
            if class_options.if_cancelled == 0:
                for student in data_objects.students:
                    if (student['name'] + ' ' + student['surname']) in group_class_list:
                        student['student_schedule'].append('Grup Dersi' + ' ' + self.tableWidget.horizontalHeaderItem(cell.column()).text()+
                                                           ' ' + self.tableWidget.verticalHeaderItem(cell.row()).text() + ' ' + item)
                for employee in data_objects.employees:
                    if (employee['name']+' '+employee['surname']) == self.tableWidget.horizontalHeaderItem(cell.column()).text():
                        employee['teacher_schedule'].append(('Grup Dersi' + ' ' + str(group_class_list) + ' ' +
                                                             self.tableWidget.verticalHeaderItem(cell.row()).text() + ' ' + item))
            else:
                pass

        elif option == "Dersi Iptal Et":
            self.check_student_attandence()
            text = self.tableWidget.item(cell.row(), cell.column()).text()
            date_temp = item.split(' ')
            date = date_temp[-1]
            time = self.tableWidget.verticalHeaderItem(cell.row()).text()
            teacher = self.tableWidget.horizontalHeaderItem(cell.column()).text()
            for student in data_objects.students:
                for schedule in student['student_schedule']:
                    if teacher in schedule and date in schedule and time in schedule:
                        cancel_class.lesson_selected_2 = schedule
            for employee in data_objects.employees:
                if teacher == employee['name'] + ' ' + employee['surname']:
                    for schedule in employee['teacher_schedule']:
                        if date in schedule and time in schedule:
                            cancel_class.lesson_selected = schedule
            cancel_class.class_cancel()

        elif option == "Yoklama Bilgisi Ekle":
            self.check_student_attandence()
            for lesson in attandence_list:
                if item in lesson and  self.tableWidget.verticalHeaderItem(cell.row()).text() in lesson and self.tableWidget.horizontalHeaderItem(cell.column()).text() in lesson:
                    check_attandence.attandence_class = lesson
                    check_attandence.check_group = cell.text()
                    check_attandence.open_attandence()
                else:
                    pass

        elif option == "Aktar (1 Hafta)":
            msg_box = QMessageBox()
            msg_box.setIcon(QMessageBox.Warning)
            msg_box.setText(
                'Ders Bir sonraki hafta ayni gun ve saate aktarilacaktir. Onayliyor musunuz?')
            msg_box.setWindowTitle('Uyarı')
            msg_box.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            msg_box.setDefaultButton(QMessageBox.Cancel)
            response = msg_box.exec()
            if response == QMessageBox.Ok:
                text = self.tableWidget.item(cell.row(), cell.column()).text()
                date_temp = item.split(' ')
                date = date_temp[-1]
                time = self.tableWidget.verticalHeaderItem(cell.row()).text()
                teacher = self.tableWidget.horizontalHeaderItem(cell.column()).text()
                new_date_str = " "
                for employee in data_objects.employees:
                    if teacher == employee['name'] + ' ' + employee['surname']:
                        for schedule in employee['teacher_schedule']:
                            if date in schedule and time in schedule:
                                date_2 = parse(date)
                                new_date = date_2 + relativedelta(weeks=1)
                                new_date_str = new_date.strftime("%Y-%m-%d")
                                new_schedule = schedule.replace(date, new_date_str)
                                employee['teacher_schedule'].append(new_schedule)
                        for schedule in employee['teacher_attended']:
                            if date in schedule and time in schedule:
                                date_2 = parse(date)
                                new_date = date_2 + relativedelta(weeks=1)
                                new_date_str = new_date.strftime("%Y-%m-%d")
                                new_schedule = schedule.replace(date, new_date_str)
                                employee['teacher_schedule'].append(new_schedule)
                        for schedule in employee['teacher_skipped']:
                            if date in schedule and time in schedule:
                                date_2 = parse(date)
                                new_date = date_2 + relativedelta(weeks=1)
                                new_date_str = new_date.strftime("%Y-%m-%d")
                                new_schedule = schedule.replace(date, new_date_str)
                                employee['teacher_schedule'].append(new_schedule)
                for student in data_objects.students:
                    for schedule in student['student_schedule']:
                        if teacher in schedule and date in schedule and time in schedule:
                            new_schedule = schedule.replace(date, new_date_str)
                            student['student_schedule'].append(new_schedule)
                    for schedule in student['student_attended']:
                        if teacher in schedule and date in schedule and time in schedule:
                            new_schedule = schedule.replace(date, new_date_str)
                            student['student_schedule'].append(new_schedule)
                    for schedule in student['student_skipped']:
                        if teacher in schedule and date in schedule and time in schedule:
                            new_schedule = schedule.replace(date, new_date_str)
                            student['student_schedule'].append(new_schedule)
            else:
                pass

        elif option == "Dersi Sil":
            msg_box = QMessageBox()
            msg_box.setIcon(QMessageBox.Warning)
            msg_box.setText(
                'Ders yoklama/iptal durumu dahil tamamen sistemden silinecektir. Onayliyor musunuz?')
            msg_box.setWindowTitle('Uyarı')
            msg_box.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            msg_box.setDefaultButton(QMessageBox.Cancel)
            response = msg_box.exec()
            if response == QMessageBox.Ok:
                text = self.tableWidget.item(cell.row(), cell.column()).text()
                date = item
                time = self.tableWidget.verticalHeaderItem(cell.row()).text()
                teacher = self.tableWidget.horizontalHeaderItem(cell.column()).text()
                for employee in data_objects.employees:
                    if teacher == employee['name'] + ' ' + employee['surname']:
                        for schedule in employee['teacher_schedule']:
                            if date in schedule and time in schedule:
                                employee['teacher_schedule'].remove(schedule)
                        for schedule in employee['teacher_attended']:
                            if date in schedule and time in schedule:
                                employee['teacher_attended'].remove(schedule)
                        for schedule in employee['teacher_skipped']:
                            if date in schedule and time in schedule:
                                employee['teacher_skipped'].remove(schedule)
                        for schedule in employee['schedule_cancelled']:
                            if date in schedule and time in schedule:
                                employee['schedule_cancelled'].remove(schedule)
                for student in data_objects.students:
                    for schedule in student['student_schedule']:
                        if teacher in schedule and date in schedule and time in schedule:
                            student['student_schedule'].remove(schedule)
                    for schedule in student['student_attended']:
                        if teacher in schedule and date in schedule and time in schedule:
                            student['student_attended'].remove(schedule)
                    for schedule in student['student_skipped']:
                        if teacher in schedule and date in schedule and time in schedule:
                            student['student_skipped'].remove(schedule)
                    for schedule in student['schedule_cancelled']:
                        if teacher in schedule and date in schedule and time in schedule:
                            student['schedule_cancelled'].remove(schedule)
            else:
                pass

        with open("student_data.txt", "w", encoding="utf-8") as f:
            f.writelines(json.dumps(data_objects.students, default=str))
        with open('student_data.txt', 'r', encoding="utf-8") as f:
            data_objects.students = json.load(f)
        with open("employee_data.txt", "w", encoding="utf-8") as f:
            f.writelines(json.dumps(data_objects.employees, default=str))
        with open('employee_data.txt', 'r', encoding="utf-8") as f:
            data_objects.employees = json.load(f)

        '''connect_database.upload_files('employee_data.txt')
        connect_database.upload_files('student_data.txt')'''

        self.load_general_schedule()


    def cellDoubleClicked(self, row, column):
        global group_class_list, current_cell, cell_clicked
        cell_widget_1 = self.tableWidget.cellWidget(row, column)
        if isinstance(cell_widget_1, (QtWidgets.QComboBox, QtWidgets.QLabel)):
            # If the widget is a QComboBox or a QLabel, ignore the double-click event
            return
        cell_widget = self.tableWidget.cellWidget(row, column)
        group_class_list = []
        class_options.open_class_options()
        if row in range(0, 10):
            item = self.tableWidget.item(0, 0).text()
        elif row in range(10,20):
            item = self.tableWidget.item(10, 0).text()
        elif row in  range(20,30):
            item = self.tableWidget.item(20, 0).text()
        elif row in  range(30,40):
            item = self.tableWidget.item(30, 0).text()
        elif row in  range(40,50):
            item = self.tableWidget.item(40, 0).text()
        elif row in  range(50,60):
            item = self.tableWidget.item(50, 0).text()
        elif row in  range(60,70):
            item = self.tableWidget.item(60, 0).text()
        else:
            pass
        for student in data_objects.students:
            for schedule in student['student_schedule']:
                if (self.tableWidget.horizontalHeaderItem(column).text() + ' ' + self.tableWidget.verticalHeaderItem(row).text() + ' ' + item) in schedule:
                    student['student_schedule'].remove(schedule)
                else:
                    pass
        if radio_button_selected == 3:
            comboBox_10 = QtWidgets.QComboBox()
            comboBox_10.addItem('                                     Grup Dersi')
            for student in group_class_list:
                comboBox_10.addItem('                                 ' + student)
            self.tableWidget.setCellWidget(row, column, comboBox_10)
            for student in data_objects.students:
                if (student['name'] + ' ' + student['surname']) in group_class_list:
                    student['student_schedule'].append('Grup Dersi' + ' ' + self.tableWidget.horizontalHeaderItem(column).text()+
                                                       ' ' + self.tableWidget.verticalHeaderItem(row).text() + ' ' + item)
            for employee in data_objects.employees:
                if (employee['name']+' '+employee['surname']) == self.tableWidget.horizontalHeaderItem(column).text():
                    employee['teacher_schedule'].append(('Grup Dersi' + ' ' + str(group_class_list) + ' ' +
                                                         self.tableWidget.verticalHeaderItem(row).text() + ' ' + item))
        if radio_button_selected == 2:
            label = QtWidgets.QLabel(group_class_list[0] + ' - Akademik Ders')
            label.setAlignment(Qt.AlignCenter)
            self.tableWidget.setCellWidget(row, column, label)
            for student in data_objects.students:
                if (student['name'] + ' ' + student['surname']) in group_class_list:
                    student['student_schedule'].append('Akademik Ders' + ' ' + self.tableWidget.horizontalHeaderItem(column).text()+
                                                       ' ' + self.tableWidget.verticalHeaderItem(row).text() + ' ' + item)
            for employee in data_objects.employees:
                if (employee['name']+' '+employee['surname']) == self.tableWidget.horizontalHeaderItem(column).text():
                    employee['teacher_schedule'].append(('Akademik Ders' + ' ' + str(group_class_list) + ' ' +
                                                         self.tableWidget.verticalHeaderItem(row).text() + ' ' + item))
        if radio_button_selected == 5:
            label = QtWidgets.QLabel(group_class_list[0] + ' - Attentioner')
            label.setAlignment(Qt.AlignCenter)
            self.tableWidget.setCellWidget(row, column, label)
            for student in data_objects.students:
                if (student['name'] + ' ' + student['surname']) in group_class_list:
                    student['student_schedule'].append(
                        'Attentioner Ders' + ' ' + self.tableWidget.horizontalHeaderItem(column).text() +
                        ' ' + self.tableWidget.verticalHeaderItem(row).text() + ' ' + item)
            for employee in data_objects.employees:
                if (employee['name'] + ' ' + employee['surname']) == self.tableWidget.horizontalHeaderItem(
                        column).text():
                    employee['teacher_schedule'].append(('Attentioner Ders' + ' ' + str(group_class_list) + ' ' +
                                                         self.tableWidget.verticalHeaderItem(row).text() + ' ' + item))
        if radio_button_selected == 0:
            pass

        with open("student_data.txt", "w", encoding="utf-8") as f:
            f.writelines(json.dumps(data_objects.students, default=str))
        with open('student_data.txt', 'r', encoding="utf-8") as f:
            data_objects.students = json.load(f)
        with open("employee_data.txt", "w", encoding="utf-8") as f:
            f.writelines(json.dumps(data_objects.employees, default=str))
        with open('employee_data.txt', 'r', encoding="utf-8") as f:
            data_objects.employees = json.load(f)

        '''connect_database.upload_files('employee_data.txt')
        connect_database.upload_files('student_data.txt')'''

        self.load_general_schedule()


    def check_student_attandence(self):
        global attandence_list, attandence_checked
        attandence_list.clear()
        for employee in data_objects.employees:
            for schedule in employee['teacher_schedule']:
                schedule_2 = schedule.split(']')
                attandence_list.append(employee['name'] + ' ' + employee['surname'] + schedule_2[-1])


    def copy_schedule_next(self):
        global group_class_list, date_list, week_1, year_1
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Warning)
        msg_box.setText('Planlanan (Tamamlanan/Iptal Olan Dersler Dahil Degildir) derslerin tam kopyasi bir sonraki haftaya aktarilaacaktir. '
                        'Bir sonraki haftada bulunan tum programlariniz tamamen silinecek ve '
                        'yerine kopyalanan haftanin verileri kaydedilecektir. Onayliyor musunuz?')
        msg_box.setWindowTitle('Uyarı')
        msg_box.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        msg_box.setDefaultButton(QMessageBox.Cancel)
        response = msg_box.exec()
        if response == QMessageBox.Ok:
            # display the QMessageBox and wait for user input
            date_list_1 = date_list
            self.step_fwd_weekly()
            date_list_2 = date_list
            self.step_back_weekly()
            temp_student_schedule = []
            temp_teacher_schedule = []
            for employee in data_objects.employees:
                for schedule in employee['teacher_schedule']:
                    if any(elem in schedule for elem in date_list_2):
                        employee['teacher_schedule'].remove(schedule)
            for student in data_objects.students:
                for schedule in student['student_schedule']:
                    if any(elem in schedule for elem in date_list_2):
                        student['student_schedule'].remove(schedule)
            for employee in data_objects.employees:
                for schedule in employee['teacher_schedule']:
                    if any(elem in schedule for elem in date_list_1):
                        index_list = [i for i, elem in enumerate(date_list_1) if elem in schedule]
                        for j in index_list:
                            new_schedule = schedule.replace(date_list_1[j], date_list_2[j])
                            employee['teacher_schedule'].append(new_schedule)


            for student in data_objects.students:
                for schedule in student['student_schedule']:
                    if any(elem in schedule for elem in date_list_1):
                        index_list = [i for i, elem in enumerate(date_list_1) if elem in schedule]
                        for j in index_list:
                            new_schedule = schedule.replace(date_list_1[j], date_list_2[j])
                            student['student_schedule'].append(new_schedule)


            with open("employee_data.txt", "w", encoding="utf-8") as f:
                f.writelines(json.dumps(data_objects.employees, default=str))
            with open("student_data.txt", "w", encoding="utf-8") as f:
                f.writelines(json.dumps(data_objects.students, default=str))
            with open('employee_data.txt', 'r', encoding="utf-8") as f:
                data_objects.employees = json.load(f)
            with open('student_data.txt', 'r', encoding="utf-8") as f:
                data_objects.students = json.load(f)
        else:
            pass
        self.step_fwd_weekly()

    def find_week_number(self):
        global week_1, year_1, firstdayofweek
        date_temp = str(self.dateEdit.date().toPyDate()).split('-')
        date_1= (int(date_temp[0]),int(date_temp[1]),int(date_temp[2]))
        date_2 = datetime.date(date_1[0],date_1[1],date_1[2])
        week_1 = date_2.isocalendar().week
        year_1 = date_2.isocalendar().year
        self.label_35.setText('{} {}. Hafta Genel Cizelge'.format(year_1,week_1))


    def step_fwd_weekly(self):
        global week_1, year_1, firstdayofweek, date_list
        self.populate_table()
        date_list = []
        if week_1 == 52:
            week_1 = 1
            year_1 = year_1 + 1
        else:
            week_1 = week_1 + 1
        firstdayofweek = datetime.date.fromisocalendar(year_1, week_1, 1)  # generating the datetime
        self.dateEdit.setCurrentSection(QtWidgets.QDateTimeEdit.DaySection)
        self.find_week_number()
        week_2 = str(firstdayofweek).split('-')
        self.dateEdit.setDate(QDate(int(week_2[0]),int(week_2[1]),int(week_2[2])))
        for i in range(7):
            self.tableWidget.setSpan(i * 10, 0, 1, len(data_objects.employees))
            self.tableWidget.setItem(i * 10, 0,
                                     QTableWidgetItem(day_of_week[i] + ' ' + str(self.dateEdit.date().toPyDate())))
            self.dateEdit.setCurrentSection(QtWidgets.QDateTimeEdit.DaySection)
            current_date = self.dateEdit.date()
            date_list.append(day_of_week[i] + ' ' + str(self.dateEdit.date().toPyDate()))
            new_date = current_date.addDays(1)
            self.dateEdit.setDate(new_date)
            font = QFont()  # Create a new font object
            font.setBold(True)  # Set the font weight to bold
            self.tableWidget.item(i * 10, 0).setFont(font)
            self.tableWidget.item(i * 10, 0).setBackground(QColor("#E6E7E7"))
            item = self.tableWidget.item(i * 10, 0)
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            item.setFlags(item.flags() & ~Qt.ItemIsEnabled)
        with open('student_data.txt', 'r', encoding="utf-8") as f:
            data_objects.students = json.load(f)
        with open('employee_data.txt', 'r', encoding="utf-8") as f:
            data_objects.employees = json.load(f)
        v_header = [' ', '9:30-10:10', '10:30-11:10', '11:30-12:10', '12:30-13:10', '13:10-14:00', '14:00-14:40',
                    '15:00-15:40', '16:00-16:40', '17:00-17:40']
        vertical_header = []
        for j in range(7):
            for i in v_header:
                vertical_header.append(i)
        self.tableWidget.setVerticalHeaderLabels(vertical_header)
        column_count = 0
        column_names = []
        for employee in data_objects.employees:
            if employee['title'] == 'Ogretmen' or  employee['title'] == 'Dil Konusmaci' or  employee['title'] == 'Psikolog':
                column_names.append((employee['name'] + ' ' + employee['surname']))
                column_count = column_count + 1
            else:
                pass
        self.tableWidget.setColumnCount(column_count)
        self.tableWidget.setHorizontalHeaderLabels(column_names)
        self.dateEdit.setDate(QDate(int(week_2[0]), int(week_2[1]), int(week_2[2])))
        self.load_general_schedule()

    def step_back_weekly(self):
        global week_1, year_1, firstdayofweek, date_list
        self.populate_table()
        date_list = []
        if week_1 == 1:
            week_1 = 52
            year_1 = year_1 - 1
        else:
            week_1 = week_1 - 1
        firstdayofweek = datetime.date.fromisocalendar(year_1, week_1, 1)  # generating the datetime
        self.dateEdit.setCurrentSection(QtWidgets.QDateTimeEdit.DaySection)
        self.find_week_number()
        week_2 = str(firstdayofweek).split('-')
        self.dateEdit.setDate(QDate(int(week_2[0]),int(week_2[1]),int(week_2[2])))
        for i in range(7):
            self.tableWidget.setSpan(i * 10, 0, 1, len(data_objects.employees))
            self.tableWidget.setItem(i * 10, 0,
                                     QTableWidgetItem(day_of_week[i] + ' ' + str(self.dateEdit.date().toPyDate())))
            self.dateEdit.setCurrentSection(QtWidgets.QDateTimeEdit.DaySection)
            current_date = self.dateEdit.date()
            date_list.append(day_of_week[i] + ' ' + str(self.dateEdit.date().toPyDate()))
            new_date = current_date.addDays(1)
            self.dateEdit.setDate(new_date)
            font = QFont()  # Create a new font object
            font.setBold(True)  # Set the font weight to bold
            self.tableWidget.item(i * 10, 0).setFont(font)
            self.tableWidget.item(i * 10, 0).setBackground(QColor("#E6E7E7"))
            item = self.tableWidget.item(i * 10, 0)
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            item.setFlags(item.flags() & ~Qt.ItemIsEnabled)
        with open('student_data.txt', 'r', encoding="utf-8") as f:
            data_objects.students = json.load(f)
        with open('employee_data.txt', 'r', encoding="utf-8") as f:
            data_objects.employees = json.load(f)
        v_header = [' ', '9:30-10:10', '10:30-11:10', '11:30-12:10', '12:30-13:10', '13:10-14:00', '14:00-14:40',
                    '15:00-15:40', '16:00-16:40', '17:00-17:40']
        vertical_header = []
        for j in range(7):
            for i in v_header:
                vertical_header.append(i)
        self.tableWidget.setVerticalHeaderLabels(vertical_header)
        column_count = 0
        column_names = []
        for employee in data_objects.employees:
            if employee['title'] == 'Ogretmen' or  employee['title'] == 'Dil Konusmaci' or  employee['title'] == 'Psikolog':
                column_names.append((employee['name'] + ' ' + employee['surname']))
                column_count = column_count + 1
            else:
                pass
        self.tableWidget.setColumnCount(column_count)
        self.tableWidget.setHorizontalHeaderLabels(column_names)
        self.dateEdit.setDate(QDate(int(week_2[0]), int(week_2[1]), int(week_2[2])))
        self.load_general_schedule()


    def combobox_top(self):
        if self.comboBox.currentText() == 'Ogrenci Verileri':
            self.frame_11.hide()
            self.frame_7.hide()
            self.frame_3.show()
        if self.comboBox.currentText() == 'Ogretmen Verileri':
            self.frame_11.hide()
            self.frame_3.hide()
            self.frame_7.show()
        if self.comboBox.currentText() == 'Personel Verileri':
            self.frame_3.hide()
            self.frame_7.hide()
            self.frame_11.show()

    def shift_data_student_right(self):
        global data_count_student
        if data_count_student == 2:
            self.frame_5.hide()
            self.frame_4.show()
            self.mark_dates_student()
            self.frame_50.hide()
            self.label_30.setText("           Ogrenci Cizelgesi")
            data_count_student = 1
            if data_objects.active_auth_level == 4:
                self.frame_5.hide()
                self.comboBox_5.hide()
                self.label_58.hide()
                self.comboBox_9.hide()
                self.label_100.hide()
                self.label_80.hide()
                self.label_49.hide()
            else:
                pass
        elif data_count_student == 0:
            self.frame_4.hide()
            self.frame_5.show()
            self.frame_50.hide()
            self.label_30.setText("  Ogrenci Bilgileri Genel Ozet")
            data_count_student = 2
            if data_objects.active_auth_level == 4:
                self.frame_5.hide()
                self.comboBox_5.hide()
                self.label_58.hide()
                self.comboBox_9.hide()
                self.label_100.hide()
                self.label_80.hide()
                self.label_49.hide()
            else:
                pass
        else:
            self.frame_4.hide()
            self.frame_5.hide()
            self.set_student_statistics()
            self.frame_50.show()
            self.label_30.setText("    Ogrenci Genel Istatistikleri")
            data_count_student = 0

    def shift_data_student_left(self):
        global data_count_student
        if data_count_student == 0:
            self.frame_5.hide()
            self.frame_4.show()
            self.mark_dates_student()
            self.frame_50.hide()
            self.label_30.setText("         Ogrenci Cizelgesi")
            data_count_student = 1
            if data_objects.active_auth_level == 4:
                self.frame_5.hide()
                self.comboBox_5.hide()
                self.label_58.hide()
                self.comboBox_9.hide()
                self.label_100.hide()
                self.label_80.hide()
                self.label_49.hide()
            else:
                pass
        elif data_count_student == 1:
            self.frame_4.hide()
            self.frame_5.show()
            self.frame_50.hide()
            self.label_30.setText("  Ogrenci Bilgileri Genel Ozet")
            data_count_student = 2
            if data_objects.active_auth_level == 4:
                self.frame_5.hide()
                self.comboBox_5.hide()
                self.label_58.hide()
                self.comboBox_9.hide()
                self.label_100.hide()
                self.label_80.hide()
                self.label_49.hide()
            else:
                pass
        else:
            self.frame_4.hide()
            self.frame_5.hide()
            self.set_student_statistics()
            self.frame_50.show()
            self.label_30.setText("    Ogrenci Genel Istatistikleri")
            data_count_student = 0

    def shift_data_teacher_right(self):
        global data_count_teacher
        if data_count_teacher == 2:
            self.frame_8.hide()
            self.frame_9.show()
            self.mark_dates_employee()
            self.label_118.hide()
            self.frame_70.hide()
            self.label_31.setText("    Ogretmen Cizelgesi")
            data_count_teacher = 1
        elif data_count_teacher == 0:
            self.frame_9.hide()
            self.frame_8.show()
            self.label_118.show()
            self.frame_70.hide()
            self.label_31.setText(" Ogretmen Bilgileri Genel Ozet")
            data_count_teacher = 2
        else:
            self.frame_8.hide()
            self.frame_9.hide()
            self.label_118.hide()
            self.set_employee_statistics()
            self.frame_70.show()
            self.label_31.setText(" Ogretmen Genel Istatisitkleri")
            data_count_teacher = 0

    def shift_data_teacher_left(self):
        global data_count_teacher
        if data_count_teacher == 0:
            self.frame_8.hide()
            self.frame_9.show()
            self.mark_dates_employee()
            self.label_118.hide()
            self.frame_70.hide()
            self.label_31.setText("           Ogretmen Cizelgesi")
            data_count_teacher = 1
        elif data_count_teacher == 1:
            self.frame_9.hide()
            self.frame_8.show()
            self.label_118.show()
            self.frame_70.hide()
            self.label_31.setText(" Ogretmen Bilgileri Genel Ozet")
            data_count_teacher = 2
        else:
            self.frame_8.hide()
            self.frame_9.hide()
            self.label_118.hide()
            self.set_employee_statistics()
            self.frame_70.show()
            self.label_31.setText(" Ogretmen Genel Istatisitkleri")
            data_count_teacher = 0

    def show_data(self):
        global date_list

        date_consistency.convert_dates_in_list(data_objects.students)
        date_consistency.convert_dates_in_list(data_objects.employees)

        with open("student_data.txt", "w", encoding="utf-8") as f:
            f.writelines(json.dumps(data_objects.students, default=str))
        with open("employee_data.txt", "w", encoding="utf-8") as f:
            f.writelines(json.dumps(data_objects.employees, default=str))

        self.tableWidget.clear()
        self.comboBox.clear()
        self.dateEdit.setDate(QDate.currentDate())
        date_list = []
        self.load_txt()
        self.find_week_number()

        with open('employee_data.txt', 'r', encoding="utf-8") as f:
            data_objects.employees = json.load(f)

        with open('student_data.txt', 'r', encoding="utf-8") as f:
            data_objects.students = json.load(f)
        self.frame_13.hide()
        self.frame_2.show()
        v_header = [' ','9:30-10:10','10:30-11:10','11:30-12:10','12:30-13:10','13:10-14:00','14:00-14:40','15:00-15:40','16:00-16:40','17:00-17:40']
        vertical_header = []
        for j in range(7):
            for i in v_header:
                vertical_header.append(i)
        self.tableWidget.setVerticalHeaderLabels(vertical_header)
        self.tableWidget.setColumnCount(len(data_objects.employees))
        column_count = 0
        column_names = []
        for employee in data_objects.employees:
            if employee['title'] == 'Ogretmen' or  employee['title'] == 'Dil Konusmaci' or  employee['title'] == 'Psikolog':
                column_names.append((employee['name'] + ' ' + employee['surname']))
                column_count = column_count + 1
            else:
                pass
        self.tableWidget.setColumnCount(column_count)
        self.tableWidget.setHorizontalHeaderLabels(column_names)
        firstdayofweek = datetime.date.fromisocalendar(year_1, week_1, 1)  # generating the datetime
        self.dateEdit.setCurrentSection(QtWidgets.QDateTimeEdit.DaySection)
        self.find_week_number()
        week_2 = str(firstdayofweek).split('-')
        self.dateEdit.setDate(QDate(int(week_2[0]), int(week_2[1]), int(week_2[2])))
        for i in range(7):
            self.tableWidget.setSpan(i * 10, 0, 1, len(data_objects.employees))
            self.tableWidget.setItem(i * 10, 0,
                                     QTableWidgetItem(day_of_week[i] + ' ' + str(self.dateEdit.date().toPyDate())))
            self.dateEdit.setCurrentSection(QtWidgets.QDateTimeEdit.DaySection)
            current_date = self.dateEdit.date()
            date_list.append(day_of_week[i] + ' ' + str(self.dateEdit.date().toPyDate()))
            new_date = current_date.addDays(1)
            self.dateEdit.setDate(new_date)
            font = QFont()  # Create a new font object
            font.setBold(True)  # Set the font weight to bold
            try:
                self.tableWidget.item(i * 10, 0).setFont(font)
                self.tableWidget.item(i * 10, 0).setBackground(QColor("#E6E7E7"))
                item = self.tableWidget.item(i * 10, 0)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                item.setFlags(item.flags() & ~Qt.ItemIsEnabled)
            except AttributeError:
                pass
        self.dateEdit.setDate(QDate.currentDate())
        self.comboBox.setCurrentIndex(1)
        self.comboBox.setCurrentIndex(0)
        self.comboBox_2.model().sort(0)
        self.load_general_schedule()
        self.pushButton_3.setDisabled(True)
        self.pushButton_4.setDisabled(True)
        self.pushButton_7.setDisabled(True)

        def process_student_data_file():
            with open('student_data.txt', 'r', encoding="utf-8") as f:
                data_objects.students = json.load(f)

            # Process the data
            for item in data_objects.students:
                name = item.get('name')
                if name and ' ' in name:
                    item['name'] = name.replace(' ', '-')
                if name.endswith('-'):
                    item['name'] = name.replace('-',
                                                '')  # Remove the trailing '-' if it is not followed by a word character

                for key, value in item.items():
                    if isinstance(value, str) and value.endswith('\n') or ('\n') in value:
                        item[key] = value.replace('\n', '')

            # Save the updated data back to the file
            with open("student_data.txt", "w", encoding="utf-8") as f:
                f.writelines(json.dumps(data_objects.students, default=str))

        process_student_data_file()

        def process_employee_data_file():
            with open('employee_data.txt', 'r', encoding="utf-8") as f:
                data_objects.employees = json.load(f)

            # Process the data
            for item in data_objects.employees:
                name = item.get('name')
                if name and ' ' in name:
                    item['name'] = name.replace(' ', '-')
                if name.endswith('-'):
                    item['name'] = name.replace('-',
                                                '')  # Remove the trailing '-' if it is not followed by a word character

                for key, value in item.items():
                    if isinstance(value, str) and value.endswith('\n') or ('\n') in value:
                        item[key] = value.replace('\n', '')

            # Save the updated data back to the file
            with open("employee_data.txt", "w", encoding="utf-8") as f:
                f.writelines(json.dumps(data_objects.employees, default=str))

        process_employee_data_file()


    def show_general_schedule(self):
        self.step_fwd_weekly()
        self.step_back_weekly()
        with open('employee_data.txt', 'r', encoding="utf-8") as f:
            data_objects.employees = json.load(f)

        with open('student_data.txt', 'r', encoding="utf-8") as f:
            data_objects.students = json.load(f)
        self.frame_2.hide()
        self.frame_13.show()
        self.pushButton_3.setEnabled(True)
        self.pushButton_4.setEnabled(True)
        self.pushButton_7.setEnabled(True)

    def log_out(self):
        msgBox = QMessageBox()
        msgBox.setWindowTitle("Cikis Yap")
        msgBox.setIcon(QMessageBox.Question)
        msgBox.setText(
            "Kullanici cikisi yapmak istiyor musunuz?")
        msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        yesButton = msgBox.button(QMessageBox.Yes)
        yesButton.setText("Evet")
        noButton = msgBox.button(QMessageBox.No)
        noButton.setText("Hayir")
        response = msgBox.exec_()
        # Perform an action based on the users response
        if response == QMessageBox.Yes:
            self.ready_to_close = True
        else:
            pass



    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Ana Sayfa"))
        self.pushButton_7.setText(_translate("Dialog", " Verileri Yonet"))
        self.pushButton_3.setText(_translate("Dialog", " Kurum Istatistikleri"))
        self.pushButton_14.setText(_translate("Dialog", " Aktar (1 Hafta)"))
        self.pushButton_8.setText(_translate("Dialog", " Cikis Yap"))
        self.pushButton_2.setText(_translate("Dialog", " Genel Program"))
        self.pushButton_4.setText(_translate("Dialog", " Yeni Ekle"))
        self.pushButton.setText(_translate("Dialog", " Verileri Goruntule"))
        self.pushButton_9.setText(_translate("Dialog", " Kaydet"))
        self.label_8.setText(_translate("Dialog", "Yonetici Ekrani"))
        self.label_30.setText(_translate("Dialog", "Ogrenci Bilgileri Genel Ozeti"))
        self.pushButton_6.setText(_translate("Dialog", ">"))
        self.pushButton_10.setText(_translate("Dialog", "<"))
        self.label_39.setText(_translate("Dialog", "Toplam Katilim: 10 Ders"))
        self.label_40.setText(_translate("Dialog", "Devamsizlik: 5 Ders"))
        self.label_41.setText(_translate("Dialog", "Durum: Aktif"))
        self.label_44.setText(_translate("Dialog", "Dogum Tarihi"))
        self.label_45.setText(_translate("Dialog", "Isim"))
        self.label_46.setText(_translate("Dialog", "Eski Ozel Egitim Kurumu"))
        self.label_47.setText(_translate("Dialog", "Soyisim"))
        self.label_48.setText(_translate("Dialog", "TC Kimlik No"))
        self.label_49.setText(_translate("Dialog", "Adres"))
        self.label_50.setText(_translate("Dialog", "Okul"))
        self.label_51.setText(_translate("Dialog", "Sinif"))
        self.label_55.setText(_translate("Dialog", "Kayit Tipi"))
        self.label_56.setText(_translate("Dialog", "Egitim Modulu"))
        self.label_57.setText(_translate("Dialog", "Servis Kullanimi"))
        self.label_58.setText(_translate("Dialog", "Saatlik Ders Ucreti"))
        self.label_59.setText(_translate("Dialog", "Kayit Tarihi"))
        self.label_60.setText(_translate("Dialog", "Baslangic Notlari ve Sorunlar"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), _translate("Dialog", "Aylik Program"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_2), _translate("Dialog", "Haftalik Program"))
        self.label_61.setText(_translate("Dialog", "Anne Isim"))
        self.label_64.setText(_translate("Dialog", "Anne Telefon No"))
        self.label_65.setText(_translate("Dialog", "Baba Telefon No"))
        self.label_68.setText(_translate("Dialog", "Baba Isim"))
        self.label_31.setText(_translate("Dialog", "Ogretmen Bilgileri Genel Ozeti"))
        self.pushButton_11.setText(_translate("Dialog", ">"))
        self.pushButton_12.setText(_translate("Dialog", "<"))
        self.label_43.setText(_translate("Dialog", "Toplam Gecmis Ders: - Ders"))
        self.label_89.setText(_translate("Dialog", "Iptal Edilen Ders: - Ders"))
        self.label_90.setText(_translate("Dialog", "Durum: Aktif"))
        self.label_91.setText(_translate("Dialog", "Dogum Tarihi"))
        self.label_92.setText(_translate("Dialog", "Isim"))
        self.label_93.setText(_translate("Dialog", "Gecmis Kurum / Kurumlar"))
        self.label_94.setText(_translate("Dialog", "Soyisim"))
        self.label_95.setText(_translate("Dialog", "TC Kimlik No"))
        self.label_97.setText(_translate("Dialog", "Mezun Oldugu Kurum & Fakulte"))
        self.label_99.setText(_translate("Dialog", "Calisma Tipi"))
        self.label_100.setText(_translate("Dialog", "Maas Bilgisi / Degisim Tarihi"))
        self.label_101.setText(_translate("Dialog", "Ise Baslangic Tarihi"))
        self.label_96.setText(_translate("Dialog", "Adres"))
        self.label_98.setText(_translate("Dialog", "Telefon"))
        self.label_119.setText(_translate("Dialog", "Mezuniyet Tarihi"))
        self.label_121.setText(_translate("Dialog", "Email"))
        self.label_130.setText(_translate("Dialog", "Calisma Gunleri:"))
        self.label_131.setText(_translate("Dialog", "Calisma Gunleri:"))
        self.tabWidget_3.setTabText(self.tabWidget_3.indexOf(self.tab_5), _translate("Dialog", "Aylik Program"))
        self.tabWidget_3.setTabText(self.tabWidget_3.indexOf(self.tab_6), _translate("Dialog", "Haftalik Program"))
        self.label_32.setText(_translate("Dialog", "Personel Bilgileri Genel Ozeti"))
        self.label_169.setText(_translate("Dialog", "Durum: Aktif"))
        self.label_170.setText(_translate("Dialog", "Dogum Tarihi"))
        self.label_171.setText(_translate("Dialog", "Isim"))
        self.label_172.setText(_translate("Dialog", "Gecmis Kurum / Kurumlar"))
        self.label_173.setText(_translate("Dialog", "Soyisim"))
        self.label_174.setText(_translate("Dialog", "TC Kimlik No"))
        self.label_175.setText(_translate("Dialog", "Mezun Oldugu Kurum & Fakulte"))
        self.label_176.setText(_translate("Dialog", "Calisma Tipi"))
        self.label_177.setText(_translate("Dialog", "Maas Bilgisi / Degisim Tarihi"))
        self.label_178.setText(_translate("Dialog", "Ise Baslangic Tarihi"))
        self.label_186.setText(_translate("Dialog", "Adres"))
        self.label_189.setText(_translate("Dialog", "Telefon"))
        self.label_190.setText(_translate("Dialog", "Mezuniyet Tarihi"))
        self.label_192.setText(_translate("Dialog", "Email"))
        self.label_194.setText(_translate("Dialog", "    Personel Bilgileri"))
        self.label_118.setText(_translate("Dialog", "    Ogretmen Bilgileri"))
        self.label_10.setText(_translate("Dialog", "Yonetici Ekrani"))
        self.pushButton_19.setText(_translate("Dialog", "<"))
        self.pushButton_20.setText(_translate("Dialog", ">"))
        self.label_35.setText(_translate("Dialog", "2023 1. Hafta Genel Cizelge"))

def open_mainpage():
    Dialog = QtWidgets.QDialog()
    Dialog.setStyle(QtWidgets.QStyleFactory.create("Fusion"))
    ui = Ui_Dialog()
    ui.setupUi(Dialog)
    Dialog.show()
    Dialog.exec_()

